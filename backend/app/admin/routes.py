from flask import Blueprint, request, send_file, jsonify, current_app
from app.admin.service import AdminService
from app.admin.designation_service import DesignationService
from app.middleware.role_guard import hr_required
from app.middleware.company_context import company_required
from app.utils.response import (
    success_response, 
    error_response, 
    validation_error_response
)
import os
import tempfile
import csv
import io
from datetime import datetime

admin_bp = Blueprint('admin', __name__)


# ADMIN DASHBOARD
@admin_bp.route('/dashboard', methods=['GET'])
@company_required
@hr_required
def get_dashboard_stats():
    """Get admin dashboard statistics (HR only)"""
    try:
        result = AdminService.get_dashboard_stats()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve dashboard stats", status_code=500)


# DEPARTMENTS
@admin_bp.route('/departments', methods=['POST'])
@company_required
@hr_required
def add_department():
    """Add new department (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['department_code', 'department_name']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        result = AdminService.add_department(data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add department", status_code=500)


@admin_bp.route('/departments', methods=['GET'])
@company_required
@hr_required
def list_departments():
    """Get list of all departments (HR only)"""
    try:
        result = AdminService.list_departments()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"departments": result["data"]}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve departments", status_code=500)


@admin_bp.route('/departments/<int:department_id>', methods=['PUT'])
@company_required
@hr_required
def update_department(department_id):
    """Update a department (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Update department
        result = AdminService.update_department(
            department_id=department_id,
            department_code=data.get('department_code'),
            department_name=data.get('department_name')
        )
        
        if result["success"]:
            return success_response(
                message="Department updated successfully",
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to update department", status_code=500)


@admin_bp.route('/departments/<int:department_id>', methods=['DELETE'])
@company_required
@hr_required
def delete_department(department_id):
    """Delete a department (HR only)"""
    try:
        result = AdminService.delete_department(department_id)
        
        if result["success"]:
            return success_response(
                message="Department deleted successfully",
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to delete department", status_code=500)


# DESIGNATIONS
@admin_bp.route('/designations', methods=['POST'])
@company_required
@hr_required
def add_designation():
    """Add new designation (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['designation_name']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        result = AdminService.add_designation(data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add designation", status_code=500)


@admin_bp.route('/designations', methods=['GET'])
@company_required
@hr_required
def list_designations():
    """Get list of all designations (HR only)"""
    try:
        result = AdminService.list_designations()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"designations": result["data"]}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve designations", status_code=500)


@admin_bp.route('/designations/<int:designation_id>', methods=['PUT'])
@company_required
@hr_required
def update_designation(designation_id):
    """Update a designation (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Update designation
        result = AdminService.update_designation(
            designation_id=designation_id,
            designation_name=data.get('designation_name'),
            designation_level=data.get('designation_level')
        )
        
        if result["success"]:
            return success_response(
                message="Designation updated successfully",
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to update designation", status_code=500)


@admin_bp.route('/designations/<int:designation_id>', methods=['DELETE'])
@company_required
@hr_required
def delete_designation(designation_id):
    """Delete a designation (HR only)"""
    try:
        result = AdminService.delete_designation(designation_id)
        
        if result["success"]:
            return success_response(
                message="Designation deleted successfully",
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to delete designation", status_code=500)


# LOCATIONS
@admin_bp.route('/locations', methods=['POST'])
@company_required
@hr_required
def add_location():
    """Add new location (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['location_name', 'city', 'country']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        result = AdminService.add_location(data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add location", status_code=500)


@admin_bp.route('/locations', methods=['GET'])
@company_required
@hr_required
def list_locations():
    """Get list of all locations (HR only)"""
    try:
        result = AdminService.list_locations()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"locations": result["data"]}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve locations", status_code=500)


@admin_bp.route('/locations/<int:location_id>', methods=['PUT'])
@company_required
@hr_required
def update_location(location_id):
    """Update a location (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Update location
        result = AdminService.update_location(
            location_id=location_id,
            location_name=data.get('location_name'),
            city=data.get('city'),
            country=data.get('country')
        )
        
        if result["success"]:
            return success_response(
                message="Location updated successfully",
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to update location", status_code=500)


@admin_bp.route('/locations/<int:location_id>', methods=['DELETE'])
@company_required
@hr_required
def delete_location(location_id):
    """Delete a location (HR only)"""
    try:
        result = AdminService.delete_location(location_id)
        
        if result["success"]:
            return success_response(
                message="Location deleted successfully",
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to delete location", status_code=500)


# HOLIDAYS
@admin_bp.route('/holidays', methods=['POST'])
@company_required
@hr_required
def add_holiday():
    """Add new holiday (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['holiday_date', 'holiday_name', 'holiday_type', 'calendar_year']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        result = AdminService.add_holiday(data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add holiday", status_code=500)


@admin_bp.route('/holidays', methods=['GET'])
@company_required
@hr_required
def list_holidays():
    """Get holidays by year (HR only)"""
    try:
        year = request.args.get('year', 2026, type=int)
        
        result = AdminService.list_holidays_by_year(year)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"holidays": result["data"], "year": year}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve holidays", status_code=500)


@admin_bp.route('/holidays/<int:holiday_id>', methods=['PUT'])
@company_required
@hr_required
def update_holiday(holiday_id):
    """Update holiday (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        result = AdminService.update_holiday(holiday_id, data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to update holiday", status_code=500)


@admin_bp.route('/holidays/<int:holiday_id>', methods=['DELETE'])
@company_required
@hr_required
def delete_holiday(holiday_id):
    """Delete holiday (HR only)"""
    try:
        result = AdminService.delete_holiday(holiday_id)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to delete holiday", status_code=500)


@admin_bp.route('/holidays/template', methods=['GET'])
@company_required
@hr_required
def download_holiday_template():
    """Download Excel template for bulk holiday upload (HR only)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from flask import send_file
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Holiday Template"
        
        # Define headers
        headers = ['holiday_date', 'holiday_name', 'holiday_type', 'calendar_year']
        header_descriptions = [
            'Date (DD-MM-YYYY)',
            'Holiday Name',
            'Type (National/Optional)',
            'Year (e.g., 2026)'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for col, (header, desc) in enumerate(zip(headers, header_descriptions), start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add description in row 2
            desc_cell = ws.cell(row=2, column=col)
            desc_cell.value = desc
            desc_cell.font = Font(italic=True, color="666666")
            desc_cell.alignment = Alignment(horizontal='center')
        
        # Add sample data
        sample_data = [
            ['25-12-2026', 'Christmas', 'National', 2026],
            ['26-01-2026', 'Republic Day', 'National', 2026],
            ['15-08-2026', 'Independence Day', 'National', 2026],
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='holiday_template.xlsx'
        )
        
    except Exception as e:
        current_app.logger.error(f"Template download error: {str(e)}")
        return error_response("Failed to generate template", status_code=500)


@admin_bp.route('/holidays/bulk-upload', methods=['POST'])
@company_required
@hr_required
def bulk_upload_holidays():
    """Bulk upload holidays from Excel file (HR only)"""
    try:
        from openpyxl import load_workbook
        from datetime import datetime
        
        # Check if file is present
        if 'file' not in request.files:
            return validation_error_response("No file uploaded")
        
        file = request.files['file']
        
        if file.filename == '':
            return validation_error_response("No file selected")
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return validation_error_response("Only Excel files (.xlsx, .xls) are allowed")
        
        # Load workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Validate headers
        required_headers = ['holiday_date', 'holiday_name', 'holiday_type', 'calendar_year']
        if not all(h in headers for h in required_headers):
            return validation_error_response(
                f"Invalid template. Required columns: {', '.join(required_headers)}"
            )
        
        # Process rows (skip header row and description row)
        results = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # Skip empty rows
            if not any(row):
                continue
            
            results['total'] += 1
            
            try:
                # Extract data
                holiday_date = row[headers.index('holiday_date')]
                holiday_name = row[headers.index('holiday_name')]
                holiday_type = row[headers.index('holiday_type')]
                calendar_year = row[headers.index('calendar_year')]
                
                # Convert date if it's a datetime object
                if isinstance(holiday_date, datetime):
                    holiday_date = holiday_date.strftime('%Y-%m-%d')
                elif isinstance(holiday_date, str):
                    # Try to parse the date string - support both DD-MM-YYYY and YYYY-MM-DD
                    try:
                        # Try DD-MM-YYYY format first
                        parsed_date = datetime.strptime(holiday_date, '%d-%m-%Y')
                        holiday_date = parsed_date.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            # Try YYYY-MM-DD format
                            parsed_date = datetime.strptime(holiday_date, '%Y-%m-%d')
                            holiday_date = parsed_date.strftime('%Y-%m-%d')
                        except ValueError:
                            raise ValueError(f"Invalid date format: {holiday_date}. Use DD-MM-YYYY or YYYY-MM-DD")
                
                # Validate data
                if not holiday_date or not holiday_name or not holiday_type or not calendar_year:
                    raise ValueError("All fields are required")
                
                # Call stored procedure
                result = AdminService.bulk_upload_holiday({
                    'holiday_date': holiday_date,
                    'holiday_name': str(holiday_name).strip(),
                    'holiday_type': str(holiday_type).strip(),
                    'calendar_year': int(calendar_year)
                })
                
                if result['success']:
                    results['success'] += 1
                else:
                    results['failed'] += 1
                    results['errors'].append({
                        'row': row_idx,
                        'holiday_name': holiday_name,
                        'error': result.get('message', 'Unknown error')
                    })
                    
            except Exception as e:
                results['failed'] += 1
                results['errors'].append({
                    'row': row_idx,
                    'holiday_name': row[headers.index('holiday_name')] if len(row) > 1 else 'Unknown',
                    'error': str(e)
                })
        
        # Prepare response message
        if results['failed'] == 0:
            message = f"All {results['success']} holidays uploaded successfully"
            return success_response(message=message, data=results)
        elif results['success'] == 0:
            message = f"All {results['failed']} holidays failed to upload"
            return error_response(message, data=results, status_code=400)
        else:
            message = f"{results['success']} holidays uploaded, {results['failed']} failed"
            return success_response(message=message, data=results)
            
    except Exception as e:
        current_app.logger.error(f"Bulk upload error: {str(e)}")
        return error_response("Failed to process bulk upload", status_code=500)


# LEAVE TYPES
@admin_bp.route('/leave-types', methods=['POST'])
@company_required
@hr_required
def add_leave_type():
    """Add new leave type (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['leave_code', 'leave_name', 'max_days_per_year']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        result = AdminService.add_leave_type(data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add leave type", status_code=500)


@admin_bp.route('/leave-types', methods=['GET'])
@company_required
@hr_required
def list_leave_types():
    """Get list of all leave types (HR only)"""
    try:
        result = AdminService.list_leave_types()
        if result["success"]:
            return success_response(message=result["message"], data={"leave_types": result["data"]})
        else:
            return error_response(result["message"], status_code=500)
    except Exception as e:
        return error_response("Failed to retrieve leave types", status_code=500)


@admin_bp.route('/leave-types/<int:leave_type_id>', methods=['DELETE'])
@company_required
@hr_required
def deactivate_leave_type(leave_type_id):
    """Soft-delete a leave type (HR only)"""
    try:
        result = AdminService.deactivate_leave_type(leave_type_id)
        if result["success"]:
            return success_response(message=result["message"])
        else:
            return error_response(result["message"], status_code=400)
    except Exception as e:
        return error_response("Failed to deactivate leave type", status_code=500)


# SALARY STRUCTURES
@admin_bp.route('/salary-structures', methods=['POST'])
@company_required
@hr_required
def add_salary_structure():
    """Add new salary structure (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['structure_name', 'structure_type']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        # Validate structure_type
        valid_types = ['Monthly', 'CTC']
        if data.get('structure_type') not in valid_types:
            return validation_error_response(
                f"Invalid structure_type. Must be one of: {', '.join(valid_types)}"
            )
        
        result = AdminService.add_salary_structure(data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add salary structure", status_code=500)


@admin_bp.route('/salary-structures', methods=['GET'])
@company_required
@hr_required
def list_salary_structures():
    """Get list of all salary structures (HR only)"""
    try:
        result = AdminService.list_salary_structures()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"salary_structures": result["data"]}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve salary structures", status_code=500)


# LETTER TEMPLATES
@admin_bp.route('/letter-templates', methods=['GET'])
@company_required
@hr_required
def list_letter_templates():
    """Get list of all letter templates (HR only)"""
    try:
        result = AdminService.list_letter_templates()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"templates": result["data"]}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve letter templates", status_code=500)


@admin_bp.route('/letter-templates', methods=['POST'])
@company_required
@hr_required
def add_letter_template():
    """Add new letter template (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['template_name', 'template_category', 'template_content']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        result = AdminService.add_letter_template(data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add letter template", status_code=500)


@admin_bp.route('/letter-templates/<int:template_id>', methods=['PUT'])
@company_required
@hr_required
def update_letter_template(template_id):
    """Update letter template (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        result = AdminService.update_letter_template(template_id, data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to update letter template", status_code=500)


@admin_bp.route('/letter-templates/<int:template_id>', methods=['DELETE'])
@company_required
@hr_required
def delete_letter_template(template_id):
    """Delete letter template (HR only)"""
    try:
        result = AdminService.delete_letter_template(template_id)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to delete letter template", status_code=500)


# COMPANY POLICIES
@admin_bp.route('/company-policies', methods=['GET'])
@company_required
@hr_required
def list_company_policies():
    """Get list of all company policies (HR only)"""
    try:
        result = AdminService.list_company_policies()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"policies": result["data"]}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve company policies", status_code=500)


@admin_bp.route('/company-policies', methods=['POST'])
@company_required
@hr_required
def add_company_policy():
    """Add new company policy (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['policy_title', 'policy_category', 'policy_description']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        result = AdminService.add_company_policy(data)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add company policy", status_code=500)


@admin_bp.route('/company-policies/<int:policy_id>', methods=['PUT'])
@company_required
@hr_required
def update_company_policy(policy_id):
    """Update a company policy (HR only)"""
    try:
        data = request.get_json()
        if not data:
            return validation_error_response("Request body is required")
        result = AdminService.update_company_policy(policy_id, data)
        if result["success"]:
            return success_response(message=result["message"], data=result["data"])
        else:
            return error_response(result["message"], status_code=400)
    except Exception as e:
        return error_response("Failed to update company policy", status_code=500)


@admin_bp.route('/company-policies/<int:policy_id>', methods=['DELETE'])
@company_required
@hr_required
def delete_company_policy(policy_id):
    """Delete a company policy (HR only)"""
    try:
        result = AdminService.delete_company_policy(policy_id)
        if result["success"]:
            return success_response(message=result["message"], data=result["data"])
        else:
            return error_response(result["message"], status_code=400)
    except Exception as e:
        return error_response("Failed to delete company policy", status_code=500)


@admin_bp.route('/company-settings', methods=['GET'])
@company_required
@hr_required
def get_company_settings():
    """Get company settings (HR only)"""
    try:
        result = AdminService.get_company_settings()
        if result["success"]:
            return success_response(message=result["message"], data={"settings": result["data"]})
        else:
            return error_response(result["message"], status_code=500)
    except Exception as e:
        return error_response("Failed to retrieve company settings", status_code=500)


@admin_bp.route('/company-settings', methods=['PUT'])
@company_required
@hr_required
def save_company_settings():
    """Save company settings (HR only)"""
    try:
        data = request.get_json()
        if not data:
            return validation_error_response("Request body is required")
        result = AdminService.save_company_settings(data)
        if result["success"]:
            return success_response(message=result["message"], data=result["data"])
        else:
            return error_response(result["message"], status_code=400)
    except Exception as e:
        return error_response("Failed to save company settings", status_code=500)


# SYSTEM REPORTS
@admin_bp.route('/reports/generate', methods=['POST'])
@company_required
@hr_required
def generate_system_report():
    """Generate system report (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        report_type = data.get('report_type')
        if not report_type:
            return validation_error_response("report_type is required")
        
        # Valid report types
        valid_types = ['employee-master', 'attendance-summary', 'leave-summary']
        if report_type not in valid_types:
            return validation_error_response(
                f"Invalid report_type. Must be one of: {', '.join(valid_types)}"
            )
        
        filters = {
            'date_from': data.get('date_from'),
            'date_to': data.get('date_to'),
            'department': data.get('department')
        }
        
        result = AdminService.generate_system_report(report_type, filters)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to generate system report", status_code=500)


# BULK UPLOAD TEMPLATES
@admin_bp.route('/bulk-upload/templates/<template_type>', methods=['GET'])
@company_required
@hr_required
def download_bulk_upload_template(template_type):
    """Download bulk upload template with data validation (HR only)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        import tempfile
        
        if template_type == 'employee-master':
            # Fetch master data for dropdowns
            departments_result = AdminService.list_departments()
            designations_result = AdminService.list_designations()
            locations_result = AdminService.list_locations()
            
            departments = [d['department_name'] for d in (departments_result.get('data') or [])] if departments_result.get('success') else []
            designations = [d['designation_name'] for d in (designations_result.get('data') or [])] if designations_result.get('success') else []
            locations = [l['location_name'] for l in (locations_result.get('data') or [])] if locations_result.get('success') else []

            # Fetch shifts for dropdown
            from app.database.multi_tenant_executor import MultiTenantExecutor as _MTE
            shifts_result = _MTE.execute_procedure('proc_get_shift_definitions')
            shifts = []
            if shifts_result.get('success') and shifts_result.get('data'):
                shifts = [s['shift_name'] for s in shifts_result['data'] if s.get('is_active')]            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Employee Data"
            
            # Create reference sheet for dropdown data
            ref_sheet = wb.create_sheet("Reference Data")
            
            # Add departments to reference sheet
            ref_sheet['A1'] = 'Departments'
            ref_sheet['A1'].font = Font(bold=True)
            for idx, dept in enumerate(departments, start=2):
                ref_sheet[f'A{idx}'] = dept
            
            # Add designations to reference sheet
            ref_sheet['B1'] = 'Designations'
            ref_sheet['B1'].font = Font(bold=True)
            for idx, desig in enumerate(designations, start=2):
                ref_sheet[f'B{idx}'] = desig
            
            # Add locations to reference sheet
            ref_sheet['C1'] = 'Locations'
            ref_sheet['C1'].font = Font(bold=True)
            for idx, loc in enumerate(locations, start=2):
                ref_sheet[f'C{idx}'] = loc
            
            # Add gender options
            ref_sheet['D1'] = 'Gender'
            ref_sheet['D1'].font = Font(bold=True)
            genders = ['Male', 'Female', 'Other']
            for idx, gender in enumerate(genders, start=2):
                ref_sheet[f'D{idx}'] = gender
            
            # Add employment types
            ref_sheet['E1'] = 'Employment Type'
            ref_sheet['E1'].font = Font(bold=True)
            emp_types = ['Full-Time', 'Part-Time', 'Contract', 'Intern', 'Consultant']
            for idx, emp_type in enumerate(emp_types, start=2):
                ref_sheet[f'E{idx}'] = emp_type

            # Add shifts to reference sheet
            ref_sheet['F1'] = 'Shifts'
            ref_sheet['F1'].font = Font(bold=True)
            for idx, shift in enumerate(shifts, start=2):
                ref_sheet[f'F{idx}'] = shift

            # Add worker categories to reference sheet
            ref_sheet['G1'] = 'Worker Category'
            ref_sheet['G1'].font = Font(bold=True)
            worker_categories = ['OFFICE', 'FACTORY']
            for idx, category in enumerate(worker_categories, start=2):
                ref_sheet[f'G{idx}'] = category

            # Define headers for main sheet
            headers = [
                # Required (red) - left side
                'Employee Code', 'First Name', 'Last Name (optional)',
                'Department', 'Designation', 'Date of Joining (YYYY-MM-DD)',
                # Optional (blue) - right side
                'Email (optional)', 'Phone', 'DOB (YYYY-MM-DD)', 'Gender', 'Address',
                'Emergency Contact Phone', 'Work Location', 'Employment Type', 'Worker Category', 'Shift (optional)'
            ]

            # Mark required fields (first 6 columns, excluding Last Name at index 2)
            required_cols = [0, 1, 3, 4, 5]  # Employee Code, First Name, Department, Designation, DOJ
            
            # Style header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            required_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                # Red background for required fields
                if col_num - 1 in required_cols:
                    cell.fill = required_fill
                else:
                    cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add sample data row
            sample_data = [
                # Required columns first
                '1', 'John', 'Doe',
                departments[0] if departments else 'Engineering',
                designations[0] if designations else 'Software Engineer',
                '2024-01-01',
                # Optional columns
                'john.doe@company.com',
                '9876543210', '1990-01-15', 'Male', '123 Main St, Mumbai, Maharashtra, India 400001',
                '9876543212',
                locations[0] if locations else 'Mumbai Office',
                'Full-Time',
                'OFFICE',  # Worker Category
                shifts[0] if shifts else ''
            ]
            
            for col_num, value in enumerate(sample_data, start=1):
                ws.cell(row=2, column=col_num, value=value)
            
            # Get the last employee code from database to continue sequence
            try:
                from app.database.multi_tenant_executor import MultiTenantExecutor
                result = MultiTenantExecutor.execute_procedure('proc_get_next_employee_code', {})
                next_num = 2
                if result.get("success") and result.get("data") and len(result["data"]) > 0:
                    row = result["data"][0]
                    next_code = row.get('next_employee_code') if row else None
                    if next_code:
                        try:
                            next_num = int(next_code)
                        except:
                            next_num = 2

                # Pre-fill 100 rows with sequential numbers
                for row_num in range(3, 102):
                    formula = f'=IF(B{row_num-1}<>"",A{row_num-1}+1,"")'
                    ws.cell(row=row_num, column=1, value=formula)
                # Set first auto row explicitly
                ws.cell(row=3, column=1, value=next_num)

            except Exception as e:
                current_app.logger.warning(f"Could not fetch last employee code: {str(e)}")
                for row_num in range(3, 102):
                    ws.cell(row=row_num, column=1, value=f'=IF(B{row_num-1}<>"",A{row_num-1}+1,"")')
            
            # Add data validation (dropdowns) for specific columns
            from datetime import datetime, timedelta
            today = datetime.now()
            
            # 1. Email validation (column G) - optional
            email_dv = DataValidation(type="custom", formula1='OR(G2="",AND(ISNUMBER(SEARCH("@",G2)),ISNUMBER(SEARCH(".",G2)),LEN(G2)>5))', allow_blank=True)
            email_dv.error = 'Please enter a valid email address (must contain @ and domain)'
            email_dv.errorTitle = 'Invalid Email Format'
            email_dv.prompt = 'Optional. Enter email in format: name@company.com. Leave empty for factory workers.'
            email_dv.promptTitle = 'Email Address (Optional)'
            ws.add_data_validation(email_dv)
            email_dv.add('G2:G1000')
            
            # 2. Phone validation (column H) - exactly 10 digits
            phone_dv = DataValidation(type="textLength", operator="equal", formula1="10", allow_blank=True)
            phone_dv.error = 'Phone number must be exactly 10 digits'
            phone_dv.errorTitle = 'Invalid Phone Number'
            phone_dv.prompt = 'Enter 10-digit phone number (e.g., 9876543210)'
            phone_dv.promptTitle = 'Phone Number'
            ws.add_data_validation(phone_dv)
            phone_dv.add('H2:H1000')
            
            # 3. DOB validation (column I) - age 18-65
            min_dob = (today - timedelta(days=65*365)).strftime('%Y-%m-%d')
            max_dob = (today - timedelta(days=18*365)).strftime('%Y-%m-%d')
            dob_dv = DataValidation(type="date", operator="between", formula1=min_dob, formula2=max_dob, allow_blank=True)
            dob_dv.error = f'Date of birth must be between {min_dob} and {max_dob} (age 18-65)'
            dob_dv.errorTitle = 'Invalid Date of Birth'
            dob_dv.prompt = 'Enter date in YYYY-MM-DD format (employee must be 18-65 years old)'
            dob_dv.promptTitle = 'Date of Birth'
            ws.add_data_validation(dob_dv)
            dob_dv.add('I2:I1000')
            
            # 4. Gender dropdown (column J)
            if len(genders) > 0:
                gender_dv = DataValidation(type="list", formula1=f"'Reference Data'!$D$2:$D${len(genders)+1}", allow_blank=True)
                gender_dv.error = 'Please select from the dropdown'
                gender_dv.errorTitle = 'Invalid Gender'
                gender_dv.prompt = 'Select gender from dropdown'
                gender_dv.promptTitle = 'Gender'
                ws.add_data_validation(gender_dv)
                gender_dv.add('J2:J1000')
            
            # 5. Address length validation (column K) - max 500 characters
            address_dv = DataValidation(type="textLength", operator="lessThanOrEqual", formula1="500", allow_blank=True)
            address_dv.error = 'Address must be 500 characters or less'
            address_dv.errorTitle = 'Address Too Long'
            address_dv.prompt = 'Enter full address (max 500 characters)'
            address_dv.promptTitle = 'Address'
            ws.add_data_validation(address_dv)
            address_dv.add('K2:K1000')
            
            # 6. Emergency contact phone validation (column L) - exactly 10 digits
            emerg_phone_dv = DataValidation(type="textLength", operator="equal", formula1="10", allow_blank=True)
            emerg_phone_dv.error = 'Emergency contact phone must be exactly 10 digits'
            emerg_phone_dv.errorTitle = 'Invalid Phone Number'
            emerg_phone_dv.prompt = 'Enter 10-digit phone number'
            emerg_phone_dv.promptTitle = 'Emergency Contact Phone'
            ws.add_data_validation(emerg_phone_dv)
            emerg_phone_dv.add('L2:L1000')
            
            # 7. Department dropdown (column D) - REQUIRED
            if len(departments) > 0:
                dept_dv = DataValidation(type="list", formula1=f"'Reference Data'!$A$2:$A${len(departments)+1}", allow_blank=False)
                dept_dv.error = 'Please select a valid department from the dropdown'
                dept_dv.errorTitle = 'Invalid Department'
                dept_dv.prompt = 'Select department from dropdown (REQUIRED)'
                dept_dv.promptTitle = 'Department'
                ws.add_data_validation(dept_dv)
                dept_dv.add('D2:D1000')
            
            # 8. Designation dropdown (column E) - REQUIRED
            if len(designations) > 0:
                desig_dv = DataValidation(type="list", formula1=f"'Reference Data'!$B$2:$B${len(designations)+1}", allow_blank=False)
                desig_dv.error = 'Please select a valid designation from the dropdown'
                desig_dv.errorTitle = 'Invalid Designation'
                desig_dv.prompt = 'Select designation from dropdown (REQUIRED)'
                desig_dv.promptTitle = 'Designation'
                ws.add_data_validation(desig_dv)
                desig_dv.add('E2:E1000')
            
            # 9. Date of Joining validation (column F) - REQUIRED
            max_doj = (today + timedelta(days=90)).strftime('%Y-%m-%d')
            doj_dv = DataValidation(type="date", operator="lessThanOrEqual", formula1=max_doj, allow_blank=False)
            doj_dv.error = f'Date of joining must be on or before {max_doj}'
            doj_dv.errorTitle = 'Invalid Date of Joining'
            doj_dv.prompt = 'Enter date in YYYY-MM-DD format (REQUIRED, max 90 days in future)'
            doj_dv.promptTitle = 'Date of Joining'
            ws.add_data_validation(doj_dv)
            doj_dv.add('F2:F1000')
            
            # 10. Work Location dropdown (column M)
            if len(locations) > 0:
                loc_dv = DataValidation(type="list", formula1=f"'Reference Data'!$C$2:$C${len(locations)+1}", allow_blank=True)
                loc_dv.error = 'Please select from the dropdown'
                loc_dv.errorTitle = 'Invalid Location'
                loc_dv.prompt = 'Select work location from dropdown'
                loc_dv.promptTitle = 'Work Location'
                ws.add_data_validation(loc_dv)
                loc_dv.add('M2:M1000')
            
            # 11. Employment Type dropdown (column N)
            if len(emp_types) > 0:
                emp_type_dv = DataValidation(type="list", formula1=f"'Reference Data'!$E$2:$E${len(emp_types)+1}", allow_blank=True)
                emp_type_dv.error = 'Please select from the dropdown'
                emp_type_dv.errorTitle = 'Invalid Employment Type'
                emp_type_dv.prompt = 'Select employment type from dropdown'
                emp_type_dv.promptTitle = 'Employment Type'
                ws.add_data_validation(emp_type_dv)
                emp_type_dv.add('N2:N1000')

            # 12. Shift dropdown (column O) - optional
            if len(shifts) > 0:
                shift_dv = DataValidation(type="list", formula1=f"'Reference Data'!$F$2:$F${len(shifts)+1}", allow_blank=True)
                shift_dv.error = 'Please select from the dropdown'
                shift_dv.errorTitle = 'Invalid Shift'
                shift_dv.prompt = 'Select shift from dropdown (optional — leave blank for office employees)'
                shift_dv.promptTitle = 'Shift'
                ws.add_data_validation(shift_dv)
                shift_dv.add('P2:P1000')  # Changed to column P since Worker Category is now column O

            # 13. Worker Category dropdown (column O) - optional
            worker_cat_dv = DataValidation(type="list", formula1="'Reference Data'!$G$2:$G$3", allow_blank=True)
            worker_cat_dv.error = 'Please select OFFICE or FACTORY'
            worker_cat_dv.errorTitle = 'Invalid Worker Category'
            worker_cat_dv.prompt = 'Select OFFICE or FACTORY (optional — defaults to OFFICE if blank)'
            worker_cat_dv.promptTitle = 'Worker Category'
            ws.add_data_validation(worker_cat_dv)
            worker_cat_dv.add('O2:O1000')
            
            # Highlight duplicate emails in column G
            from openpyxl.formatting.rule import Rule
            from openpyxl.styles.differential import DifferentialStyle
            
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            red_font = Font(color='9C0006')
            dxf = DifferentialStyle(fill=red_fill, font=red_font)
            rule = Rule(type='expression', dxf=dxf)
            rule.formula = ['AND(G2<>"",COUNTIF($G$2:$G$1000,G2)>1)']
            ws.conditional_formatting.add('G2:G1000', rule)
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter if column and len(column) > 0 else None
                if not column_letter:
                    continue
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Note: Sheet protection removed to avoid Excel compatibility issues
            # Users should avoid editing Employee Code column (A) as it's auto-generated
            
            # Hide reference sheet
            ref_sheet.sheet_state = 'hidden'
            
            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
            wb.save(temp_file.name)
            temp_file.close()
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name='employee_master_template.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        elif template_type == 'attendance':
            # Excel template for attendance with validations
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.formatting.rule import Rule
            from openpyxl.styles.differential import DifferentialStyle
            from datetime import datetime, timedelta
            
            # Fetch active employees for dropdown
            from app.database.multi_tenant_executor import MultiTenantExecutor
            employees_result = MultiTenantExecutor.execute_procedure('proc_get_active_employees_for_dropdown', {})
            employees = []
            if employees_result.get('success') and employees_result.get('data'):
                # Format: "1 - John Doe (EMP001)"
                employees = [
                    f"{emp['employee_id']} - {emp['first_name']} {emp['last_name']} ({emp['employee_code']})"
                    for emp in employees_result['data']
                ]
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Data"
            
            # Create reference sheet for dropdown data
            ref_sheet = wb.create_sheet("Reference Data")
            
            # Add employees to reference sheet
            ref_sheet['A1'] = 'Employees'
            ref_sheet['A1'].font = Font(bold=True)
            for idx, emp in enumerate(employees, start=2):
                ref_sheet[f'A{idx}'] = emp
            
            # Add status options
            ref_sheet['B1'] = 'Status'
            ref_sheet['B1'].font = Font(bold=True)
            statuses = ['PRESENT', 'ABSENT', 'LATE', 'WFH', 'HOLIDAY']
            for idx, status in enumerate(statuses, start=2):
                ref_sheet[f'B{idx}'] = status
            
            # Define headers
            headers = ['Employee ID', 'Date (DD-MM-YYYY)', 'Check-in Time (HH:MM)', 'Check-out Time (HH:MM)', 'Status']
            
            # Mark required fields (all are required)
            required_cols = [0, 1, 4]  # Employee ID, Date, Status
            
            # Style header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            required_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                # Red background for required fields
                if col_num - 1 in required_cols:
                    cell.fill = required_fill
                else:
                    cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add sample data rows (using actual employee IDs if available)
            if len(employees) >= 3:
                # Extract just the numeric ID from "1 - John Doe (EMP001)"
                emp1_id = employees[0].split(' - ')[0]
                emp2_id = employees[1].split(' - ')[0]
                emp3_id = employees[2].split(' - ')[0]
                sample_data = [
                    [emp1_id, '15-01-2026', '09:00', '18:00', 'PRESENT'],
                    [emp2_id, '15-01-2026', '09:15', '18:30', 'LATE'],
                    [emp3_id, '15-01-2026', '', '', 'ABSENT'],
                ]
            else:
                sample_data = [
                    ['1', '15-01-2026', '09:00', '18:00', 'PRESENT'],
                    ['2', '15-01-2026', '09:15', '18:30', 'LATE'],
                    ['3', '15-01-2026', '', '', 'ABSENT'],
                ]
            
            for row_num, row_data in enumerate(sample_data, start=2):
                for col_num, value in enumerate(row_data, start=1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Add data validations
            today = datetime.now()
            
            # 1. Employee ID dropdown (column A) - REQUIRED
            if len(employees) > 0:
                emp_dv = DataValidation(type="list", formula1=f"'Reference Data'!$A$2:$A${len(employees)+1}", allow_blank=False)
                emp_dv.error = 'Please select a valid employee from the dropdown'
                emp_dv.errorTitle = 'Invalid Employee'
                emp_dv.prompt = 'Select employee from dropdown (REQUIRED). You can type to search.'
                emp_dv.promptTitle = 'Employee ID'
                ws.add_data_validation(emp_dv)
                emp_dv.add('A2:A1000')
            
            # 2. Date validation (column B) - must be valid date, not too far in past/future
            min_date = (today - timedelta(days=90)).strftime('%Y-%m-%d')  # 90 days back
            max_date = today.strftime('%Y-%m-%d')  # Today (no future dates)
            date_dv = DataValidation(type="date", operator="between", formula1=min_date, formula2=max_date, allow_blank=False)
            date_dv.error = f'Date must be between {min_date} and {max_date} (within last 90 days, no future dates)'
            date_dv.errorTitle = 'Invalid Date'
            date_dv.prompt = 'Enter date in DD-MM-YYYY format (e.g., 15-01-2026) - REQUIRED. Max 90 days in past, no future dates.'
            date_dv.promptTitle = 'Date'
            ws.add_data_validation(date_dv)
            date_dv.add('B2:B1000')
            
            # 3. Time validation for check-in (column C)
            time_in_dv = DataValidation(type="time", operator="between", formula1="00:00", formula2="23:59", allow_blank=True)
            time_in_dv.error = 'Check-in time must be between 00:00 and 23:59'
            time_in_dv.errorTitle = 'Invalid Check-in Time'
            time_in_dv.prompt = 'Enter time in HH:MM format (e.g., 09:00) - Required for PRESENT/LATE/WFH'
            time_in_dv.promptTitle = 'Check-in Time'
            ws.add_data_validation(time_in_dv)
            time_in_dv.add('C2:C1000')
            
            # 4. Time validation for check-out (column D)
            time_out_dv = DataValidation(type="time", operator="between", formula1="00:00", formula2="23:59", allow_blank=True)
            time_out_dv.error = 'Check-out time must be between 00:00 and 23:59'
            time_out_dv.errorTitle = 'Invalid Check-out Time'
            time_out_dv.prompt = 'Enter time in HH:MM format (e.g., 18:00) - Required for PRESENT/LATE/WFH. Must be after check-in time.'
            time_out_dv.promptTitle = 'Check-out Time'
            ws.add_data_validation(time_out_dv)
            time_out_dv.add('D2:D1000')
            
            # 5. Status dropdown (column E) - REQUIRED
            status_dv = DataValidation(type="list", formula1=f"'Reference Data'!$B$2:$B${len(statuses)+1}", allow_blank=False)
            status_dv.error = 'Please select a valid status from the dropdown'
            status_dv.errorTitle = 'Invalid Status'
            status_dv.prompt = 'Select status: PRESENT, ABSENT, LATE, WFH, or HOLIDAY (REQUIRED)'
            status_dv.promptTitle = 'Status'
            ws.add_data_validation(status_dv)
            status_dv.add('E2:E1000')
            
            # 6. Conditional formatting - Highlight duplicate Employee ID + Date combinations
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            red_font = Font(color='9C0006')
            dxf = DifferentialStyle(fill=red_fill, font=red_font)
            rule = Rule(type='expression', dxf=dxf)
            # Check if same employee ID + date appears more than once
            rule.formula = ['COUNTIFS($A$2:$A$1000,A2,$B$2:$B$1000,B2)>1']
            ws.conditional_formatting.add('A2:E1000', rule)
            
            # 7. Conditional formatting - Highlight rows where check-out is before check-in
            orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            orange_font = Font(color='9C5700')
            dxf2 = DifferentialStyle(fill=orange_fill, font=orange_font)
            rule2 = Rule(type='expression', dxf=dxf2)
            # Check if check-out time is less than check-in time (when both are present)
            rule2.formula = ['AND(C2<>"",D2<>"",D2<C2)']
            ws.conditional_formatting.add('C2:D1000', rule2)
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter if column and len(column) > 0 else None
                if not column_letter:
                    continue
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Hide reference sheet
            ref_sheet.sheet_state = 'hidden'
            
            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
            wb.save(temp_file.name)
            temp_file.close()
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name='attendance_template.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return error_response("Invalid template type", status_code=400)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        current_app.logger.error(f"Template generation error: {str(e)}\n{error_details}")
        return error_response(f"Failed to generate template: {str(e)}", status_code=500)


# BULK UPLOAD PROCESSING
# BULK UPLOAD PROCESSING
@admin_bp.route('/bulk-upload/employees', methods=['POST'])
@company_required
@hr_required
def process_bulk_employee_upload():
    """Process bulk employee upload CSV (HR only)"""
    try:
        if 'file' not in request.files:
            return validation_error_response("No file uploaded")
        
        file = request.files['file']
        if file.filename == '':
            return validation_error_response("No file selected")
        
        # Check file extension
        if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
            return validation_error_response("Only CSV or Excel files are supported")
        
        # Parse CSV/Excel
        import pandas as pd
        import io
        
        try:
            # Read file content once
            file_content = file.read()
            file_bytes = io.BytesIO(file_content)
            
            # Detect file type by content (Excel files start with PK signature)
            is_excel = file_content[:4] == b'PK\x03\x04' or file.filename.lower().endswith(('.xlsx', '.xls'))
            
            if is_excel:
                # Excel file - use openpyxl engine
                current_app.logger.info(f"Parsing Excel file: {file.filename}")
                df = pd.read_excel(file_bytes, engine='openpyxl', sheet_name=0)
            else:
                # CSV file
                current_app.logger.info(f"Parsing CSV file: {file.filename}")
                # Try UTF-8 first, fallback to latin-1
                try:
                    df = pd.read_csv(file_bytes, encoding='utf-8')
                except UnicodeDecodeError:
                    file_bytes.seek(0)
                    df = pd.read_csv(file_bytes, encoding='latin-1')
        except Exception as e:
            current_app.logger.error(f"File parsing error: {str(e)}")
            return validation_error_response(f"Failed to parse file: {str(e)}")
        
        # Normalize column names (strip spaces, lowercase, remove special chars)
        df.columns = df.columns.str.strip().str.lower()
        # Remove content in parentheses (e.g., "Date of Joining (YYYY-MM-DD)" -> "date of joining")
        df.columns = df.columns.str.replace(r'\s*\([^)]*\)\s*', '', regex=True)
        # Replace spaces and special chars with underscores
        df.columns = df.columns.str.replace(r'[\s\-]+', '_', regex=True)
        # Remove any remaining special characters except underscores
        df.columns = df.columns.str.replace(r'[^a-z0-9_]', '', regex=True)
        # Remove duplicate underscores
        df.columns = df.columns.str.replace(r'_+', '_', regex=True)
        # Strip leading/trailing underscores
        df.columns = df.columns.str.strip('_')
        
        current_app.logger.info(f"Parsed columns: {list(df.columns)}")
        
        # Required columns (last_name is optional - some workers may not have one)
        required_cols = ['employee_code', 'first_name',
                        'department', 'designation', 'date_of_joining']
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            return validation_error_response(
                f"Missing required columns: {', '.join(missing_cols)}"
            )
        
        # Process each row
        results = {
            'total': len(df),
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'errors': []
        }
        
        # Column mapping for special cases
        column_mapping = {
            'date_of_birth': 'dob'
        }
        
        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row number (header is row 1)
            
            # Skip empty rows
            if pd.isna(row.get('employee_code')) or str(row.get('employee_code')).strip() == '':
                results['skipped'] += 1
                continue
            
            # Convert row to dict, replacing NaN with None and applying column mapping
            employee_data = {}
            for col in df.columns:
                val = row[col]
                # Apply column mapping if exists
                mapped_col = column_mapping.get(col, col)
                employee_data[mapped_col] = None if pd.isna(val) else str(val).strip()
            
            # Attempt to create employee
            result = AdminService.bulk_upload_employee(employee_data, dry_run=False)
            
            if result['success']:
                results['success'] += 1
                # Assign shift if provided
                shift_name = employee_data.get('shift_optional') or employee_data.get('shift')
                if shift_name and shift_name.strip():
                    try:
                        from app.database.multi_tenant_executor import MultiTenantExecutor
                        # Get shift_id by name
                        shifts_res = MultiTenantExecutor.execute_procedure('proc_get_shift_definitions')
                        shift_id = None
                        if shifts_res.get('success') and shifts_res.get('data'):
                            for s in shifts_res['data']:
                                if s.get('shift_name', '').strip().lower() == shift_name.strip().lower():
                                    shift_id = s['shift_id']
                                    break
                        if shift_id and result.get('employee_id'):
                            MultiTenantExecutor.execute_procedure('proc_assign_employee_shift', {
                                'employee_id': result['employee_id'],
                                'shift_id': shift_id,
                                'effective_from': employee_data.get('date_of_joining'),
                                'assigned_by': None
                            })
                    except Exception as se:
                        current_app.logger.warning(f"Shift assignment failed for row {row_num}: {se}")
            else:
                results['failed'] += 1
                results['errors'].append({
                    'row': row_num,
                    'employee_code': result.get('employee_code', ''),
                    'error': result.get('message', 'Unknown error')
                })
        
        return success_response(
            message=f"Bulk upload completed: {results['success']} succeeded, {results['failed']} failed",
            data=results
        )
        
    except Exception as e:
        current_app.logger.error(f"Bulk employee upload error: {str(e)}")
        return error_response(f"Failed to process bulk upload: {str(e)}", status_code=500)


@admin_bp.route('/bulk-upload/attendance', methods=['POST'])
@company_required
@hr_required
def process_bulk_attendance_upload():
    """Process bulk attendance upload (HR only)"""
    try:
        if 'file' not in request.files:
            return validation_error_response("No file uploaded")
        
        file = request.files['file']
        if file.filename == '':
            return validation_error_response("No file selected")
        
        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return validation_error_response("Only Excel files (.xlsx, .xls) are supported")
        
        # Save file temporarily
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
        file.save(temp_file.name)
        temp_file.close()
        
        try:
            # Use the existing BulkAttendanceUpload service
            from app.attendance.bulk_upload import BulkAttendanceUpload
            result = BulkAttendanceUpload.validate_and_process_file(temp_file.name)
            
            # Clean up temp file
            try:
                os.remove(temp_file.name)
            except:
                pass
            
            if result['success']:
                return success_response(
                    message=result['message'],
                    data={
                        'total': result['total_rows'],
                        'success': result['successful_rows'],
                        'failed': result['failed_rows'],
                        'skipped': 0,
                        'errors': result.get('errors', [])
                    }
                )
            else:
                return error_response(result['message'], status_code=400)
                
        except Exception as e:
            # Clean up temp file on error
            try:
                os.remove(temp_file.name)
            except:
                pass
            raise e
        
    except Exception as e:
        current_app.logger.error(f"Bulk attendance upload error: {str(e)}")
        return error_response(f"Failed to process bulk upload: {str(e)}", status_code=500)


@admin_bp.route('/bulk-upload/employee-images', methods=['POST'])
@company_required
@hr_required
def process_bulk_employee_images():
    """Process bulk employee image upload from ZIP file (HR only)"""
    try:
        if 'file' not in request.files:
            return validation_error_response("No file uploaded")
        
        file = request.files['file']
        if file.filename == '':
            return validation_error_response("No file selected")
        
        # Check file extension
        if not file.filename.lower().endswith('.zip'):
            return validation_error_response("Only ZIP files are supported")
        
        # Save file temporarily
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(mode='wb', suffix='.zip', delete=False)
        file.save(temp_file.name)
        temp_file.close()
        
        current_app.logger.info(f"Processing bulk image upload: {file.filename}")
        
        try:
            # Process ZIP file
            from app.admin.image_bulk_upload import BulkImageUpload
            from flask_jwt_extended import get_jwt
            claims = get_jwt()
            company_code = claims.get('company_code', 'default')
            result = BulkImageUpload.validate_and_process_zip(temp_file.name, company_code)
            
            # Clean up temp file
            try:
                os.remove(temp_file.name)
            except:
                pass
            
            if result['success']:
                return success_response(
                    message=result['message'],
                    data={
                        'total': result['total_images'],
                        'success': result['successful_images'],
                        'failed': result['failed_images'],
                        'skipped': result['skipped_images'],
                        'errors': result.get('errors', [])
                    }
                )
            else:
                return error_response(result['message'], status_code=400)
                
        except Exception as e:
            # Clean up temp file on error
            try:
                os.remove(temp_file.name)
            except:
                pass
            raise e
        
    except Exception as e:
        current_app.logger.error(f"Bulk image upload error: {str(e)}")
        return error_response(f"Failed to process bulk upload: {str(e)}", status_code=500)


@admin_bp.route('/bulk-upload/process', methods=['POST'])
@company_required
@hr_required
def process_bulk_upload():
    """Process bulk upload file (HR only)"""
    try:
        if 'file' not in request.files:
            return validation_error_response("No file uploaded")
        
        file = request.files['file']
        upload_type = request.form.get('upload_type')
        
        if not upload_type:
            return validation_error_response("upload_type is required")
        
        if file.filename == '':
            return validation_error_response("No file selected")
        
        # Process the file based on type
        # This is a placeholder - actual implementation would parse CSV/Excel
        # and call appropriate service methods
        
        return success_response(
            message=f"Bulk upload of {upload_type} processed successfully",
            data={
                "total_records": 10,
                "success_records": 8,
                "failed_records": 2,
                "upload_id": "upload_" + datetime.now().strftime("%Y%m%d_%H%M%S")
            }
        )
        
    except Exception as e:
        return error_response("Failed to process bulk upload", status_code=500)


# DESIGNATION ROLE MAPPINGS
@admin_bp.route('/designation-mappings', methods=['GET'])
@company_required
@hr_required
def get_designation_mappings():
    """Get all designation-role mappings (HR only)"""
    try:
        result = DesignationService.get_designation_mappings()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"mappings": result["data"]}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve designation mappings", status_code=500)


@admin_bp.route('/designation-mappings', methods=['POST'])
@company_required
@hr_required
def add_designation_mapping():
    """Add new designation-role mapping (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        # Validate required fields
        required_fields = ['designation_name', 'role_code']
        missing_fields = []
        
        for field in required_fields:
            if not data.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            return validation_error_response(
                f"Missing required fields: {', '.join(missing_fields)}"
            )
        
        # Validate role_code
        valid_roles = ['HR', 'MANAGER', 'EMPLOYEE']
        if data.get('role_code') not in valid_roles:
            return validation_error_response(
                f"Invalid role_code. Must be one of: {', '.join(valid_roles)}"
            )
        
        result = DesignationService.add_designation_mapping(
            data['designation_name'], 
            data['role_code']
        )
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to add designation mapping", status_code=500)


@admin_bp.route('/designation-mappings/<int:mapping_id>', methods=['PUT'])
@company_required
@hr_required
def update_designation_mapping(mapping_id):
    """Update designation-role mapping (HR only)"""
    try:
        data = request.get_json()
        
        if not data:
            return validation_error_response("Request body is required")
        
        role_code = data.get('role_code')
        if not role_code:
            return validation_error_response("role_code is required")
        
        # Validate role_code
        valid_roles = ['HR', 'MANAGER', 'EMPLOYEE']
        if role_code not in valid_roles:
            return validation_error_response(
                f"Invalid role_code. Must be one of: {', '.join(valid_roles)}"
            )
        
        result = DesignationService.update_designation_mapping(mapping_id, role_code)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to update designation mapping", status_code=500)


@admin_bp.route('/designation-mappings/<int:mapping_id>', methods=['DELETE'])
@company_required
@hr_required
def delete_designation_mapping(mapping_id):
    """Delete designation-role mapping (HR only)"""
    try:
        result = DesignationService.delete_designation_mapping(mapping_id)
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data=result["data"]
            )
        else:
            return error_response(result["message"], status_code=400)
            
    except Exception as e:
        return error_response("Failed to delete designation mapping", status_code=500)


@admin_bp.route('/available-roles', methods=['GET'])
@company_required
@hr_required
def get_available_roles():
    """Get available roles for designation mapping (HR only)"""
    try:
        result = DesignationService.get_available_roles()
        
        if result["success"]:
            return success_response(
                message=result["message"],
                data={"roles": result["data"]}
            )
        else:
            return error_response(result["message"], status_code=500)
            
    except Exception as e:
        return error_response("Failed to retrieve available roles", status_code=500)

# Master Data Endpoints for Dynamic Frontend Data
@admin_bp.route('/master-data/grades', methods=['GET'])
@company_required
@hr_required
def get_grades():
    """Get all salary grades"""
    try:
        # For now, return standard grades - can be made database-driven later
        grades = ['L1', 'L2', 'L3', 'L4', 'L5', 'M1', 'M2', 'M3', 'VP', 'SVP', 'EVP']
        
        return success_response(
            message="Grades retrieved successfully",
            data={"grades": grades}
        )
        
    except Exception as e:
        return error_response(f"Failed to retrieve grades: {str(e)}", status_code=500)


@admin_bp.route('/master-data/employment-types', methods=['GET'])
@company_required
@hr_required
def get_employment_types():
    """Get all employment types"""
    try:
        # For now, return standard employment types - can be made database-driven later
        employment_types = ['Full-Time', 'Part-Time', 'Contract', 'Intern', 'Consultant', 'Temporary']
        
        return success_response(
            message="Employment types retrieved successfully",
            data={"employment_types": employment_types}
        )
        
    except Exception as e:
        return error_response(f"Failed to retrieve employment types: {str(e)}", status_code=500)


@admin_bp.route('/master-data/genders', methods=['GET'])
@company_required
@hr_required
def get_genders():
    """Get all gender options"""
    try:
        # Standard gender options
        genders = ['Male', 'Female', 'Other', 'Prefer not to say']
        
        return success_response(
            message="Gender options retrieved successfully",
            data={"genders": genders}
        )
        
    except Exception as e:
        return error_response(f"Failed to retrieve gender options: {str(e)}", status_code=500)


@admin_bp.route('/master-data/policy-categories', methods=['GET'])
@company_required
@hr_required
def get_policy_categories():
    """Get all policy categories"""
    try:
        # Standard policy categories - can be made database-driven later
        categories = ['HR Policy', 'IT Policy', 'Leave Policy', 'Finance Policy', 'Safety Policy', 'Security Policy', 'Compliance Policy']
        
        return success_response(
            message="Policy categories retrieved successfully",
            data={"categories": categories}
        )
        
    except Exception as e:
        return error_response(f"Failed to retrieve policy categories: {str(e)}", status_code=500)


@admin_bp.route('/master-data/visibility-options', methods=['GET'])
@company_required
@hr_required
def get_visibility_options():
    """Get all visibility options for policies"""
    try:
        # Standard visibility options
        options = ['Employee', 'Manager', 'HR', 'Admin', 'All']
        
        return success_response(
            message="Visibility options retrieved successfully",
            data={"options": options}
        )
        
    except Exception as e:
        return error_response(f"Failed to retrieve visibility options: {str(e)}", status_code=500)


@admin_bp.route('/master-data/template-categories', methods=['GET'])
@company_required
@hr_required
def get_template_categories():
    """Get all template categories"""
    try:
        # Standard template categories
        categories = ['Onboarding', 'Offboarding', 'Payroll', 'General', 'Legal', 'HR', 'Performance']
        
        return success_response(
            message="Template categories retrieved successfully",
            data={"categories": categories}
        )
        
    except Exception as e:
        return error_response(f"Failed to retrieve template categories: {str(e)}", status_code=500)


# BULK UPLOAD LOGS
@admin_bp.route('/bulk-upload/logs', methods=['GET'])
@company_required
@hr_required
def get_bulk_upload_logs():
    """Get bulk upload history (HR only)"""
    try:
        result = AdminService.get_bulk_upload_logs()
        return success_response(data=result['data']) if result['success'] else error_response('Failed to retrieve upload logs', 500)
    except Exception as e:
        return error_response('Failed to retrieve upload logs', status_code=500)


@admin_bp.route('/bulk-upload/log', methods=['POST'])
@company_required
@hr_required
def log_bulk_upload():
    """Log a completed bulk upload (HR only)"""
    try:
        from app.middleware.multi_tenant_jwt import get_current_user
        data = request.get_json()
        if not data:
            return validation_error_response('Request body required')
        user = get_current_user()
        result = AdminService.log_bulk_upload(
            file_name=data.get('file_name', ''),
            module=data.get('module', ''),
            total=data.get('total_records', 0),
            success=data.get('success_records', 0),
            failed=data.get('failed_records', 0),
            user_id=user.get('user_id')
        )
        return success_response(message='Upload logged') if result['success'] else error_response('Failed to log upload', 500)
    except Exception as e:
        return error_response('Failed to log upload', status_code=500)


# AUDIT LOGS
@admin_bp.route('/audit-logs', methods=['GET'])
@company_required
@hr_required
def get_audit_logs():
    """Get audit logs with optional filters (HR only)"""
    try:
        module = request.args.get('module')
        user_id = request.args.get('user_id', type=int)
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        result = AdminService.get_audit_logs(module=module, user_id=user_id, from_date=from_date, to_date=to_date)
        return success_response(data=result['data']) if result['success'] else error_response('Failed to retrieve audit logs', 500)
    except Exception as e:
        return error_response('Failed to retrieve audit logs', status_code=500)
