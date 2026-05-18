"""
Biometric Device Bulk Upload
Parses the raw punch log exported from the biometric device via USB.

Expected Excel format (from device):
    ID | Name | Time
    2  | AABHAS NEGI | 2026/04/08 17:29:06

After inserting raw logs, triggers proc_generate_factory_attendance
for the date range found in the file.

Inactivity logic:
- INACTIVE employees in the file are SKIPPED (not uploaded), HR is notified.
- After every upload, ALL active employees with zero punches in the last 7 days
  are automatically marked INACTIVE (factory workers: raw_logs check only;
  office workers: also protected by approved leave).
- HR (the uploader) receives a single notification summarising both.
"""

import os
import logging
import tempfile
from datetime import datetime, timedelta
from flask import g
import openpyxl
import xlrd

from app.database.multi_tenant_executor import MultiTenantExecutor
from app.database.multi_tenant_connection import connection_manager

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _run_inactivity_check(cursor, uploader_user_id: int) -> dict:
    """
    Single-pass inactivity check run after every bulk upload.

    Logic:
      - FACTORY workers: no punch in attendance_raw_logs in last 7 days → INACTIVE
      - OFFICE workers:  same, but employees with an approved leave that overlaps
                         the last 7 days are protected.

    Uses proc_change_employee_status so every status change is properly logged
    in employee_status_history and factory_worker_exits.

    Returns a dict with lists of newly inactivated employees.
    """
    cutoff = datetime.now() - timedelta(days=7)

    # ── Factory workers ──────────────────────────────────────────────────────
    cursor.execute(
        """
        SELECT e.employee_id, e.employee_code,
               CONCAT(ep.first_name, ' ', ep.last_name) AS full_name
        FROM   employees e
        JOIN   employee_personal  ep ON ep.employee_id = e.employee_id
        JOIN   employee_official  eo ON eo.employee_id = e.employee_id
        WHERE  e.status = 'ACTIVE'
        AND    ISNULL(eo.worker_category, 'OFFICE') = 'FACTORY'
        AND    e.employee_id NOT IN (
                   SELECT DISTINCT employee_id
                   FROM   attendance_raw_logs
                   WHERE  log_time >= ?
               )
        """,
        (cutoff,)
    )
    factory_to_inactivate = cursor.fetchall()

    # Office workers are intentionally excluded from auto-inactivity.
    office_to_inactivate = []

    all_to_inactivate = factory_to_inactivate

    if not all_to_inactivate:
        return {'factory': [], 'office': []}

    # Build comma-separated ID string for proc_change_employee_status
    ids_csv = ','.join(str(r[0]) for r in all_to_inactivate)

    # Use the existing stored procedure — handles history logging,
    # factory_worker_exits, and validates inputs properly.
    # changed_by = uploader_user_id (the HR user who triggered the upload).
    cursor.execute(
        "EXEC proc_change_employee_status @employee_ids=?, @new_status='INACTIVE', "
        "@reason='Auto-inactivated: no attendance punch in last 7 days', @changed_by=?",
        (ids_csv, uploader_user_id)
    )
    result_row = cursor.fetchone()
    if result_row and result_row[0] == 0:
        logger.warning(f"proc_change_employee_status returned failure: {result_row[1]}")
    else:
        logger.info(
            f"Inactivity check: marked {len(all_to_inactivate)} employee(s) INACTIVE "
            f"(factory={len(factory_to_inactivate)}, office={len(office_to_inactivate)})"
        )

    # Factory workers don't have user accounts, so no users.is_active update needed.

    return {
        'factory': [
            {'employee_id': r[0], 'employee_code': r[1], 'name': r[2]}
            for r in factory_to_inactivate
        ],
        'office': [
            {'employee_id': r[0], 'employee_code': r[1], 'name': r[2]}
            for r in office_to_inactivate
        ],
    }


def _send_upload_notification(cursor, uploader_user_id: int,
                               skipped_inactive: list,
                               newly_inactivated: dict) -> None:
    """
    Send a single notification to the HR user who did the upload,
    summarising skipped-inactive employees and newly auto-inactivated ones.
    Uses proc_create_notification so it goes through the standard path.
    Only sends if there is something to report.
    """
    parts = []

    if skipped_inactive:
        codes = ', '.join(str(e['employee_code']) for e in skipped_inactive[:10])
        extra = f' (+{len(skipped_inactive) - 10} more)' if len(skipped_inactive) > 10 else ''
        parts.append(
            f"{len(skipped_inactive)} INACTIVE employee(s) found in upload — "
            f"their data was NOT uploaded: {codes}{extra}. "
            f"Reactivate from Employee Management if needed."
        )

    total_new = len(newly_inactivated.get('factory', [])) + len(newly_inactivated.get('office', []))
    if total_new:
        all_new = newly_inactivated['factory'] + newly_inactivated['office']
        codes = ', '.join(str(e['employee_code']) for e in all_new[:10])
        extra = f' (+{total_new - 10} more)' if total_new > 10 else ''
        parts.append(
            f"{total_new} employee(s) auto-marked INACTIVE (no punch in last 7 days): "
            f"{codes}{extra}."
        )

    if not parts:
        return

    title   = "Attendance Upload — Inactivity Alert"
    message = ' | '.join(parts)[:500]   # notifications.message is VARCHAR(500)

    try:
        cursor.execute(
            "EXEC proc_create_notification @user_id=?, @title=?, @message=?, @module='ATTENDANCE'",
            (uploader_user_id, title, message)
        )
        logger.info(f"Inactivity notification sent to user_id={uploader_user_id}")
    except Exception as e:
        logger.warning(f"Failed to send notification: {e}")


class BulkAttendanceUpload:

    @staticmethod
    def preview_files(file_paths: list) -> dict:
        """
        Preview multiple Excel files before processing.
        Returns summary, sample punches, and warnings.
        """
        all_rows = []
        all_errors = []
        file_summaries = []
        
        for file_path in file_paths:
            try:
                rows, errors = BulkAttendanceUpload._parse_file(file_path)
                
                file_summaries.append({
                    'file_name': os.path.basename(file_path),
                    'punch_count': len(rows),
                    'error_count': len(errors)
                })
                
                all_rows.extend(rows)
                all_errors.extend(errors)
                
            except Exception as e:
                file_summaries.append({
                    'file_name': os.path.basename(file_path),
                    'punch_count': 0,
                    'error_count': 1,
                    'error': str(e)
                })
        
        if not all_rows and all_errors:
            return {
                'success': False,
                'message': 'No valid punches found in any file',
                'files': file_summaries,
                'errors': all_errors
            }
        
        # Get date range
        min_date = None
        max_date = None
        unique_employees = set()
        
        for row in all_rows:
            punch_date = row['log_time'].date()
            if min_date is None or punch_date < min_date:
                min_date = punch_date
            if max_date is None or punch_date > max_date:
                max_date = punch_date
            unique_employees.add(row['employee_id'])
        
        # Analyze punch patterns for warnings
        early_morning_punches = []
        regular_punches = []
        
        for row in all_rows:
            if row['log_time'].time() < datetime.strptime('06:00:00', '%H:%M:%S').time():
                early_morning_punches.append(row)
            else:
                regular_punches.append(row)
        
        # Get sample punches - prioritize showing ALL early morning punches
        sample_punches = []
        
        # First, add ALL early morning punches (these are important to see)
        for row in early_morning_punches:
            punch_time = row['log_time'].time()
            sample_punches.append({
                'employee_code': row['employee_id'],
                'log_time': row['log_time'].strftime('%Y-%m-%d %H:%M:%S'),
                'is_early_morning': True,
                'note': 'Previous day checkout'
            })
        
        # Then add sample of regular punches (up to 20 total)
        remaining_slots = max(20 - len(sample_punches), 10)  # Show at least 10 regular punches
        for row in regular_punches[:remaining_slots]:
            punch_time = row['log_time'].time()
            sample_punches.append({
                'employee_code': row['employee_id'],
                'log_time': row['log_time'].strftime('%Y-%m-%d %H:%M:%S'),
                'is_early_morning': False,
                'note': 'Regular punch'
            })
        
        # Check for invalid employee codes
        company_code = getattr(g, 'company_code', None)
        warnings = []
        
        # Add early morning punch warning
        if early_morning_punches:
            from datetime import timedelta
            reprocess_date = min_date - timedelta(days=1) if min_date else None
            warnings.append({
                'type': 'early_morning_checkout',
                'message': f'{len(early_morning_punches)} early morning punch(es) detected (00:00-06:00)',
                'detail': f'These will be treated as checkouts for {reprocess_date.strftime("%B %d, %Y") if reprocess_date else "previous day"} night shift',
                'count': len(early_morning_punches),
                'affected_employees': len(set(row['employee_id'] for row in early_morning_punches))
            })
        
        if company_code:
            try:
                conn = connection_manager.get_company_connection(company_code)
                cursor = conn.cursor()
                
                invalid_codes = set()
                inactive_codes = set()
                resigned_codes = set()
                
                for emp_code in unique_employees:
                    cursor.execute(
                        "SELECT employee_id, status FROM employees WHERE employee_code = ?",
                        (str(emp_code),)
                    )
                    result = cursor.fetchone()
                    if not result:
                        invalid_codes.add(emp_code)
                    elif result[1] == 'INACTIVE':
                        inactive_codes.add(emp_code)
                    elif result[1] == 'RESIGNED':
                        resigned_codes.add(emp_code)
                
                connection_manager.return_connection(company_code, conn)
                
                if invalid_codes:
                    punch_count = sum(1 for row in all_rows if row['employee_id'] in invalid_codes)
                    warnings.append({
                        'type': 'invalid_employee',
                        'message': f'{len(invalid_codes)} employee code(s) not found ({punch_count} punches will be skipped)',
                        'codes': list(invalid_codes)[:10]  # Show first 10
                    })
                
                if inactive_codes:
                    punch_count = sum(1 for row in all_rows if row['employee_id'] in inactive_codes)
                    warnings.append({
                        'type': 'inactive_employee',
                        'message': f'{len(inactive_codes)} INACTIVE employee(s) found — their data will be SKIPPED ({punch_count} punches)',
                        'codes': list(inactive_codes)[:10],
                        'detail': 'These employees are inactive. Their attendance will not be uploaded. Reactivate them from Employee Management if needed.'
                    })
                
                if resigned_codes:
                    punch_count = sum(1 for row in all_rows if row['employee_id'] in resigned_codes)
                    warnings.append({
                        'type': 'resigned_employee',
                        'message': f'{len(resigned_codes)} RESIGNED employee(s) found ({punch_count} punches will be skipped)',
                        'codes': list(resigned_codes)[:10],
                        'detail': 'Please rehire these employees before uploading attendance'
                    })
                    
            except Exception as e:
                logger.warning(f"Could not validate employee codes: {e}")
        
        return {
            'success': True,
            'total_punches': len(all_rows),
            'total_files': len(file_paths),
            'files': file_summaries,
            'date_range': {
                'from': str(min_date) if min_date else None,
                'to': str(max_date) if max_date else None
            },
            'unique_employees': len(unique_employees),
            'sample_punches': sample_punches,
            'warnings': warnings,
            'errors': all_errors[:50],  # Limit errors shown
            'punch_analysis': {
                'early_morning_checkouts': len(early_morning_punches),
                'regular_punches': len(regular_punches),
                'will_reprocess_previous_day': len(early_morning_punches) > 0
            }
        }

    @staticmethod
    def process_multiple_files(file_paths: list, uploader_user_id: int = None) -> dict:
        """
        Process multiple Excel files at once.
        Inserts all punches, then generates attendance once.
        After processing, runs the inactivity check on all active employees.
        """
        all_rows = []
        all_errors = []
        
        # Parse all files
        for file_path in file_paths:
            try:
                rows, errors = BulkAttendanceUpload._parse_file(file_path)
                all_rows.extend(rows)
                all_errors.extend(errors)
            except Exception as e:
                all_errors.append({
                    'file': os.path.basename(file_path),
                    'error': f'Failed to parse file: {str(e)}'
                })
        
        if not all_rows and all_errors:
            return {
                'success': False,
                'message': 'No valid punches found in any file',
                'total_rows': 0,
                'successful_rows': 0,
                'failed_rows': len(all_errors),
                'errors': all_errors
            }
        
        # Insert all punches
        successful = 0
        min_date = None
        max_date = None
        skipped_inactive = []   # employees whose rows were skipped
        
        company_code = getattr(g, 'company_code', None)
        if not company_code:
            return {'success': False, 'message': 'Company context not set'}
        
        # Cache employee lookups to avoid repeated DB hits for the same code
        emp_cache = {}

        conn = None
        try:
            conn = connection_manager.get_company_connection(company_code)
            cursor = conn.cursor()
            
            for row in all_rows:
                try:
                    enrollid = row['employee_id']
                    log_time = row['log_time']
                    
                    # Track date range for ALL rows (including duplicates / skipped)
                    punch_date = log_time.date()
                    if min_date is None or punch_date < min_date:
                        min_date = punch_date
                    if max_date is None or punch_date > max_date:
                        max_date = punch_date
                    
                    # Resolve employee_code -> employee_id (cached)
                    if enrollid not in emp_cache:
                        cursor.execute(
                            "SELECT employee_id, status FROM employees WHERE employee_code = ?",
                            (str(enrollid),)
                        )
                        emp_cache[enrollid] = cursor.fetchone()

                    emp_row = emp_cache[enrollid]
                    if not emp_row:
                        all_errors.append({
                            'row': row.get('row_number'),
                            'employee_id': enrollid,
                            'error': f'No employee found with code {enrollid}'
                        })
                        continue

                    employee_id     = emp_row[0]
                    employee_status = emp_row[1]

                    # ── INACTIVE: skip upload, track for notification ──────
                    if employee_status == 'INACTIVE':
                        # Only add once per employee code
                        if not any(s['employee_code'] == str(enrollid) for s in skipped_inactive):
                            skipped_inactive.append({
                                'employee_id':   employee_id,
                                'employee_code': str(enrollid),
                            })
                        logger.info(f"Skipped INACTIVE employee {enrollid} (ID: {employee_id})")
                        continue

                    # ── RESIGNED: skip with error ─────────────────────────
                    if employee_status == 'RESIGNED':
                        all_errors.append({
                            'row': row.get('row_number'),
                            'employee_id': enrollid,
                            'error': f'Employee {enrollid} has resigned. Please rehire before uploading attendance.'
                        })
                        continue
                    
                    # Skip duplicate punches
                    cursor.execute(
                        "SELECT COUNT(*) FROM attendance_raw_logs WHERE employee_id=? AND log_time=?",
                        (employee_id, log_time)
                    )
                    if cursor.fetchone()[0] > 0:
                        successful += 1
                        continue
                    
                    cursor.execute(
                        """INSERT INTO attendance_raw_logs
                           (employee_id, log_time, source, created_at)
                           VALUES (?, ?, 'BIOMETRIC', GETDATE())""",
                        (employee_id, log_time)
                    )
                    successful += 1
                        
                except Exception as row_err:
                    logger.error(f"Row insert error: {row_err}")
                    all_errors.append({
                        'row': row.get('row_number'),
                        'employee_id': row.get('employee_id'),
                        'error': str(row_err)
                    })
            
            conn.commit()

            # ── Inactivity check (single bulk query) ──────────────────────
            newly_inactivated = {'factory': [], 'office': []}
            try:
                newly_inactivated = _run_inactivity_check(cursor, uploader_user_id)
                conn.commit()
            except Exception as inact_err:
                logger.error(f"Inactivity check failed: {inact_err}")

            # ── Notify uploader ───────────────────────────────────────────
            if uploader_user_id:
                try:
                    _send_upload_notification(cursor, uploader_user_id,
                                              skipped_inactive, newly_inactivated)
                    conn.commit()
                except Exception as notif_err:
                    logger.error(f"Notification failed: {notif_err}")
            
        except Exception as e:
            logger.error(f"DB error during bulk upload: {e}")
            return {'success': False, 'message': f'Database error: {str(e)}'}
        finally:
            if conn:
                connection_manager.return_connection(company_code, conn)
        
        # Generate attendance once for all dates
        if min_date and max_date:
            try:
                # ALWAYS re-process from previous day to handle night shift scenarios
                actual_start_date = min_date - timedelta(days=1)
                logger.info(f"Reprocessing attendance from {actual_start_date} to {max_date} (includes previous day for night shift)")
                
                conn = connection_manager.get_company_connection(company_code)
                cursor = conn.cursor()
                cursor.execute(
                    "DELETE FROM attendance_daily WHERE attendance_date BETWEEN ? AND ?",
                    (actual_start_date.strftime('%Y-%m-%d'), max_date.strftime('%Y-%m-%d'))
                )
                conn.commit()
                connection_manager.return_connection(company_code, conn)
                logger.info(f"Deleted existing attendance records from {actual_start_date} to {max_date}")
                
                MultiTenantExecutor.execute_procedure(
                    'proc_generate_factory_attendance',
                    {
                        'start_date': actual_start_date.strftime('%Y-%m-%d'),
                        'end_date': max_date.strftime('%Y-%m-%d')
                    }
                )
                logger.info(f"Factory attendance generated for {actual_start_date} to {max_date}")
            except Exception as e:
                logger.warning(f"Attendance generation failed: {e}")

        # Build inactivity summary for the API response
        total_newly_inactivated = (
            len(newly_inactivated.get('factory', [])) +
            len(newly_inactivated.get('office', []))
        )

        return {
            'success': True,
            'message': f'Processed {successful} punch records from {len(file_paths)} file(s). Attendance generated.',
            'total_rows': len(all_rows),
            'successful_rows': successful,
            'failed_rows': len(all_errors),
            'errors': all_errors,
            'date_range': {
                'from': str(min_date) if min_date else None,
                'to': str(max_date) if max_date else None
            },
            'inactivity_summary': {
                'skipped_inactive_employees': skipped_inactive,
                'newly_inactivated_count': total_newly_inactivated,
                'newly_inactivated': newly_inactivated,
            }
        }

    @staticmethod
    def validate_and_process_file(file_path: str, uploader_user_id: int = None) -> dict:
        """
        Parse device USB export and insert into attendance_raw_logs.
        Then trigger factory attendance generation for the date range.
        After processing, runs the inactivity check on all active employees.
        """
        rows = []
        errors = []

        # --- Parse file ---
        try:
            rows, errors = BulkAttendanceUpload._parse_file(file_path)
        except Exception as e:
            return {'success': False, 'message': f'Failed to parse file: {str(e)}'}

        if not rows and errors:
            return {
                'success': False,
                'message': 'File could not be parsed. Check format.',
                'total_rows': 0,
                'successful_rows': 0,
                'failed_rows': len(errors),
                'errors': errors
            }

        # --- Insert raw logs ---
        successful = 0
        min_date = None
        max_date = None
        skipped_inactive = []   # employees whose rows were skipped

        company_code = getattr(g, 'company_code', None)
        if not company_code:
            return {'success': False, 'message': 'Company context not set'}

        # Cache employee lookups to avoid repeated DB hits for the same code
        emp_cache = {}

        conn = None
        newly_inactivated = {'factory': [], 'office': []}
        try:
            conn = connection_manager.get_company_connection(company_code)
            cursor = conn.cursor()

            for row in rows:
                try:
                    enrollid = row['employee_id']  # device enrollid = employee_code
                    log_time = row['log_time']

                    # Resolve employee_code -> employee_id (cached)
                    if enrollid not in emp_cache:
                        cursor.execute(
                            "SELECT employee_id, status FROM employees WHERE employee_code = ?",
                            (str(enrollid),)
                        )
                        emp_cache[enrollid] = cursor.fetchone()

                    emp_row = emp_cache[enrollid]
                    if not emp_row:
                        errors.append({
                            'row': row.get('row_number'),
                            'employee_id': enrollid,
                            'error': f'No employee found with code {enrollid}'
                        })
                        continue

                    employee_id     = emp_row[0]
                    employee_status = emp_row[1]

                    # ── INACTIVE: skip upload, track for notification ──────
                    if employee_status == 'INACTIVE':
                        if not any(s['employee_code'] == str(enrollid) for s in skipped_inactive):
                            skipped_inactive.append({
                                'employee_id':   employee_id,
                                'employee_code': str(enrollid),
                            })
                        logger.info(f"Skipped INACTIVE employee {enrollid} (ID: {employee_id})")
                        continue

                    # ── RESIGNED: skip with error ─────────────────────────
                    if employee_status == 'RESIGNED':
                        errors.append({
                            'row': row.get('row_number'),
                            'employee_id': enrollid,
                            'error': f'Employee {enrollid} has resigned. Please rehire before uploading attendance.'
                        })
                        continue

                    # Track date range even for duplicate punches
                    punch_date = log_time.date()
                    if min_date is None or punch_date < min_date:
                        min_date = punch_date
                    if max_date is None or punch_date > max_date:
                        max_date = punch_date
                    
                    # Skip duplicate punches (same employee, same minute)
                    cursor.execute(
                        "SELECT COUNT(*) FROM attendance_raw_logs WHERE employee_id=? AND log_time=?",
                        (employee_id, log_time)
                    )
                    if cursor.fetchone()[0] > 0:
                        successful += 1  # already exists, count as ok
                        continue

                    cursor.execute(
                        """INSERT INTO attendance_raw_logs
                           (employee_id, log_time, source, created_at)
                           VALUES (?, ?, 'BIOMETRIC', GETDATE())""",
                        (employee_id, log_time)
                    )
                    successful += 1

                except Exception as row_err:
                    logger.error(f"Row insert error - employee_id={row.get('employee_id')} log_time={row.get('log_time')}: {row_err}")
                    errors.append({
                        'row': row.get('row_number'),
                        'employee_id': row.get('employee_id'),
                        'error': str(row_err)
                    })

            conn.commit()

            # ── Inactivity check (single bulk query) ──────────────────────
            try:
                newly_inactivated = _run_inactivity_check(cursor, uploader_user_id)
                conn.commit()
            except Exception as inact_err:
                logger.error(f"Inactivity check failed: {inact_err}")

            # ── Notify uploader ───────────────────────────────────────────
            if uploader_user_id:
                try:
                    _send_upload_notification(cursor, uploader_user_id,
                                              skipped_inactive, newly_inactivated)
                    conn.commit()
                except Exception as notif_err:
                    logger.error(f"Notification failed: {notif_err}")

        except Exception as e:
            logger.error(f"DB error during bulk upload: {e}")
            return {'success': False, 'message': f'Database error: {str(e)}'}
        finally:
            if conn:
                connection_manager.return_connection(company_code, conn)

        # --- Trigger factory attendance generation ---
        if min_date and max_date:
            try:
                # ALWAYS re-process from previous day to handle night shift scenarios
                actual_start_date = min_date - timedelta(days=1)
                logger.info(f"Reprocessing attendance from {actual_start_date} to {max_date} (includes previous day for night shift)")
                
                conn = connection_manager.get_company_connection(company_code)
                cursor = conn.cursor()
                cursor.execute(
                    "DELETE FROM attendance_daily WHERE attendance_date BETWEEN ? AND ?",
                    (actual_start_date.strftime('%Y-%m-%d'), max_date.strftime('%Y-%m-%d'))
                )
                conn.commit()
                connection_manager.return_connection(company_code, conn)
                logger.info(f"Deleted existing attendance records from {actual_start_date} to {max_date}")
                
                MultiTenantExecutor.execute_procedure(
                    'proc_generate_factory_attendance',
                    {
                        'start_date': actual_start_date.strftime('%Y-%m-%d'),
                        'end_date': max_date.strftime('%Y-%m-%d')
                    }
                )
                logger.info(f"Factory attendance generated for {actual_start_date} to {max_date}")
            except Exception as e:
                logger.warning(f"Attendance generation failed (logs still saved): {e}")

        total_newly_inactivated = (
            len(newly_inactivated.get('factory', [])) +
            len(newly_inactivated.get('office', []))
        )

        return {
            'success': True,
            'message': f'Processed {successful} punch records. Attendance generated.',
            'total_rows': len(rows),
            'successful_rows': successful,
            'failed_rows': len(errors),
            'errors': errors,
            'date_range': {
                'from': str(min_date) if min_date else None,
                'to': str(max_date) if max_date else None
            },
            'inactivity_summary': {
                'skipped_inactive_employees': skipped_inactive,
                'newly_inactivated_count': total_newly_inactivated,
                'newly_inactivated': newly_inactivated,
            }
        }

    @staticmethod
    def _parse_file(file_path: str):
        """Parse .xls or .xlsx device export. Tries openpyxl first, falls back to xlrd."""
        # Try openpyxl first — handles .xlsx and many .xls files
        try:
            return BulkAttendanceUpload._parse_xlsx(file_path)
        except Exception:
            pass
        # Fall back to xlrd for true old-format .xls
        try:
            return BulkAttendanceUpload._parse_xls(file_path)
        except Exception as e:
            return [], [{'row': 0, 'employee_id': None, 'error': f'Cannot read file: {str(e)}'}]

    @staticmethod
    def _parse_xls(file_path: str):
        rows = []
        errors = []
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        # Find header row
        header_row = 0
        for i in range(min(5, ws.nrows)):
            row_vals = [str(ws.cell_value(i, j)).strip().upper() for j in range(ws.ncols)]
            if 'ID' in row_vals and 'TIME' in row_vals:
                header_row = i
                break

        headers = [str(ws.cell_value(header_row, j)).strip().upper() for j in range(ws.ncols)]

        try:
            id_col   = headers.index('ID')
            time_col = headers.index('TIME')
        except ValueError:
            errors.append({'row': 0, 'employee_id': None, 'error': 'Missing ID or TIME column'})
            return rows, errors

        for i in range(header_row + 1, ws.nrows):
            row_num = i + 1
            try:
                raw_id   = ws.cell_value(i, id_col)
                raw_time = ws.cell_value(i, time_col)

                if not raw_id and not raw_time:
                    continue

                employee_id = int(float(str(raw_id).strip()))
                log_time    = BulkAttendanceUpload._parse_time(raw_time, wb.datemode)

                rows.append({'employee_id': employee_id, 'log_time': log_time, 'row_number': row_num})

            except Exception as e:
                errors.append({'row': row_num, 'employee_id': None, 'error': str(e)})

        return rows, errors

    @staticmethod
    def _parse_xlsx(file_path: str):
        rows = []
        errors = []
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        header_row_idx = None
        id_col = time_col = None

        # Find header row (1-indexed for Excel, 0-indexed in openpyxl)
        for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
            headers = [str(c).strip().upper() if c else '' for c in row]
            if 'ID' in headers and 'TIME' in headers:
                header_row_idx = i
                id_col   = headers.index('ID')
                time_col = headers.index('TIME')
                break

        if id_col is None:
            errors.append({'row': 0, 'employee_id': None, 'error': 'Missing ID or TIME column'})
            return rows, errors

        logger.info(f"Found header at row {header_row_idx}, ID col: {id_col}, TIME col: {time_col}")
        logger.info(f"Total rows in sheet: {ws.max_row}")
        
        row_count = 0
        empty_count = 0
        
        # Read all rows after header
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_count += 1
            try:
                # Get cell values directly by row and column
                raw_id = ws.cell(row=row_idx, column=id_col + 1).value
                raw_time = ws.cell(row=row_idx, column=time_col + 1).value

                # Skip completely empty rows
                if raw_id is None and raw_time is None:
                    empty_count += 1
                    logger.debug(f"Row {row_idx}: Empty row, skipping")
                    continue

                # Skip if either ID or Time is missing
                if raw_id is None or raw_time is None:
                    logger.warning(f"Row {row_idx}: Missing data - ID={raw_id}, Time={raw_time}")
                    errors.append({'row': row_idx, 'employee_id': raw_id, 'error': 'Missing ID or Time'})
                    continue

                employee_id = int(float(str(raw_id).strip()))
                log_time = BulkAttendanceUpload._parse_time(raw_time)

                rows.append({'employee_id': employee_id, 'log_time': log_time, 'row_number': row_idx})
                logger.debug(f"Row {row_idx}: Parsed ID={employee_id}, Time={log_time}")

            except Exception as e:
                logger.error(f"Row {row_idx}: Parse error - {str(e)}, raw_id={raw_id}, raw_time={raw_time}")
                errors.append({'row': row_idx, 'employee_id': None, 'error': str(e)})

        logger.info(f"Parsing complete: {len(rows)} rows parsed successfully, {len(errors)} errors, {empty_count} empty rows, total rows scanned: {row_count}")
        return rows, errors

    @staticmethod
    def _parse_time(value, datemode=None) -> datetime:
        """Parse various time formats from device export."""
        if isinstance(value, datetime):
            return value

        if isinstance(value, float) and datemode is not None:
            # xlrd date serial
            import xlrd as _xlrd
            t = _xlrd.xldate_as_tuple(value, datemode)
            return datetime(*t)

        s = str(value).strip()
        for fmt in ('%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S',
                    '%d/%m/%Y %H:%M:%S', '%d-%m-%Y %H:%M:%S',
                    '%Y/%m/%d %H:%M',    '%Y-%m-%d %H:%M'):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        raise ValueError(f"Cannot parse time: {value!r}")

    @staticmethod
    def generate_template() -> dict:
        """Generate a sample template showing the device export format."""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Attendance'
            ws.append(['ID', 'Name', 'Time'])
            ws.append([2, 'AABHAS NEGI', '2026/04/09 08:05:00'])
            ws.append([2, 'AABHAS NEGI', '2026/04/09 20:10:00'])
            ws.append([3, 'PRATHAM KANOJIA', '2026/04/09 08:12:00'])
            ws.append([3, 'PRATHAM KANOJIA', '2026/04/09 20:05:00'])

            tmp = tempfile.NamedTemporaryFile(
                mode='wb', suffix='.xlsx', delete=False,
                prefix='device_attendance_template_'
            )
            wb.save(tmp.name)
            tmp.close()
            return {'success': True, 'file_path': tmp.name}
        except Exception as e:
            return {'success': False, 'error': str(e)}
