"""
Factory Payroll Routes
API endpoints for factory worker rate management and payroll
"""
from flask import request, jsonify, current_app
from flask_jwt_extended import get_jwt_identity, get_jwt
from datetime import datetime
from app.factory_payroll import factory_payroll_bp
from app.factory_payroll.service import FactoryPayrollService
from app.middleware.jwt_required import jwt_required
from app.middleware.role_guard import role_required
from app.middleware.company_context import company_required
from app.utils.response import success_response, error_response


# ============================================
# RATE MANAGEMENT ENDPOINTS
# ============================================

@factory_payroll_bp.route('/rates/assign', methods=['POST'])
@jwt_required
@company_required
@role_required("HR")
def assign_worker_rate():
    """
    Assign or update daily rate for a factory worker
    
    Request Body:
        {
            "employee_id": 218,
            "daily_rate": 575.00,
            "effective_from": "2026-04-01"
        }
    """
    try:
        data = request.get_json()
        
        if not data:
            return error_response("Request body is required", 400)
        
        employee_id = data.get('employee_id')
        daily_rate = data.get('daily_rate')
        effective_from = data.get('effective_from')
        
        # Validation
        if not employee_id:
            return error_response("employee_id is required", 400)
        
        if not daily_rate:
            return error_response("daily_rate is required", 400)
        
        if daily_rate <= 0:
            return error_response("daily_rate must be greater than zero", 400)
        
        if not effective_from:
            return error_response("effective_from date is required", 400)
        
        # Get current user ID
        created_by = get_jwt_identity()
        
        # Call service
        result = FactoryPayrollService.assign_worker_rate(
            employee_id=employee_id,
            daily_rate=daily_rate,
            effective_from=effective_from,
            created_by=created_by
        )
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
            
    except Exception as e:
        return error_response(f"Error assigning rate: {str(e)}", 500)


@factory_payroll_bp.route('/rates/current/<int:employee_id>', methods=['GET'])
@jwt_required
@company_required
@role_required("HR", "MANAGER")
def get_worker_current_rate(employee_id):
    """
    Get current active rate for a factory worker
    
    Query Parameters:
        as_of_date (optional): Date to check rate (YYYY-MM-DD)
    """
    try:
        as_of_date = request.args.get('as_of_date')
        
        result = FactoryPayrollService.get_worker_current_rate(
            employee_id=employee_id,
            as_of_date=as_of_date
        )
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 404)
            
    except Exception as e:
        return error_response(f"Error retrieving rate: {str(e)}", 500)


@factory_payroll_bp.route('/rates/history/<int:employee_id>', methods=['GET'])
@jwt_required
@company_required
@role_required("HR", "MANAGER")
def get_worker_rate_history(employee_id):
    """
    Get complete rate history for a factory worker
    """
    try:
        result = FactoryPayrollService.get_worker_rate_history(employee_id)
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 404)
            
    except Exception as e:
        return error_response(f"Error retrieving rate history: {str(e)}", 500)


@factory_payroll_bp.route('/rates/bulk-assign', methods=['POST'])
@jwt_required
@company_required
@role_required("HR")
def bulk_assign_rates():
    """
    Bulk assign rates — each entry can have its own effective_from date.
    Request Body:
        {
            "rates": [
                {"employee_id": 214, "daily_rate": 500.00, "effective_from": "2026-05-01"},
                {"employee_id": 215, "daily_rate": 550.00, "effective_from": "2026-04-01"}
            ]
        }
    effective_from is optional per entry — defaults to today if omitted.
    """
    try:
        data = request.get_json()
        if not data:
            return error_response("Request body is required", 400)

        rates = data.get('rates')
        if not rates or not isinstance(rates, list) or len(rates) == 0:
            return error_response("rates array is required and cannot be empty", 400)

        today = datetime.now().strftime('%Y-%m-%d')

        # Validate each entry and fill default effective_from
        for idx, entry in enumerate(rates):
            if 'employee_id' not in entry:
                return error_response(f"rates[{idx}]: employee_id is required", 400)
            if 'daily_rate' not in entry:
                return error_response(f"rates[{idx}]: daily_rate is required", 400)
            if entry['daily_rate'] <= 0:
                return error_response(f"rates[{idx}]: daily_rate must be greater than zero", 400)
            # Default effective_from to today if not provided
            if not entry.get('effective_from'):
                entry['effective_from'] = today

        created_by = get_jwt_identity()

        result = FactoryPayrollService.bulk_assign_rates(
            rate_data_list=rates,
            created_by=created_by
        )

        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)

    except Exception as e:
        return error_response(f"Error in bulk rate assignment: {str(e)}", 500)


@factory_payroll_bp.route('/workers-with-rates', methods=['GET'])
@jwt_required
@company_required
@role_required("HR", "MANAGER")
def get_all_workers_with_rates():
    """Get all factory workers with their current rates"""
    try:
        employee_status = request.args.get('employee_status', 'ACTIVE')
        result = FactoryPayrollService.get_all_workers_with_rates(employee_status)
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
            
    except Exception as e:
        return error_response(f"Error retrieving factory workers: {str(e)}", 500)


# ============================================
# RATE UPLOAD TEMPLATE
# ============================================

@factory_payroll_bp.route('/rates/template', methods=['GET'])
@jwt_required
@company_required
@role_required("HR")
def download_rate_template():
    """Download pre-filled Excel template for bulk rate upload"""
    try:
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        import tempfile

        # Fetch all active factory workers with current rates
        result = FactoryPayrollService.get_all_workers_with_rates('ACTIVE')
        workers = result.get('data', []) if result.get('success') else []

        wb = Workbook()
        ws = wb.active
        ws.title = 'Rate Upload'

        # ── Colour palette (matches employee master) ──────────────────────
        BLUE_DARK   = '1F3864'   # title bar
        BLUE_MID    = '2E75B6'   # required header
        BLUE_LIGHT  = 'BDD7EE'   # info header (read-only)
        GREEN_DARK  = '375623'   # editable header
        GREEN_LIGHT = 'E2EFDA'   # editable data rows (alternating)
        GREY_LIGHT  = 'F2F2F2'   # read-only data rows (alternating)
        WHITE       = 'FFFFFF'
        YELLOW      = 'FFFF00'   # highlight editable column header

        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr(color_hex):
            return PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')

        def cell_style(cell, bg, font_color='000000', bold=False, center=False, wrap=False):
            cell.fill = hdr(bg)
            cell.font = Font(color=font_color, bold=bold, size=10)
            cell.alignment = Alignment(
                horizontal='center' if center else 'left',
                vertical='center',
                wrap_text=wrap
            )
            cell.border = border

        # ── Row 1: Title bar ──────────────────────────────────────────────
        ws.merge_cells('A1:F1')
        title = ws['A1']
        title.value = 'FACTORY WORKER DAILY RATE — BULK UPLOAD TEMPLATE'
        title.fill = hdr(BLUE_DARK)
        title.font = Font(color=WHITE, bold=True, size=13)
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # ── Row 2: Sub-title / date ───────────────────────────────────────
        from datetime import date
        ws.merge_cells('A2:F2')
        sub = ws['A2']
        sub.value = f'Generated: {date.today().strftime("%d %b %Y")}  |  Fill ONLY the "New Daily Rate" column  |  Leave blank to skip that employee'
        sub.fill = hdr(BLUE_LIGHT)
        sub.font = Font(color=BLUE_DARK, italic=True, size=9)
        sub.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

        # ── Row 3: blank spacer ───────────────────────────────────────────
        ws.row_dimensions[3].height = 6

        # ── Row 4: Column headers ─────────────────────────────────────────
        headers = [
            ('Employee\nCode',        'A', BLUE_MID,   WHITE, 12),
            ('Employee Name',         'B', BLUE_LIGHT, BLUE_DARK, 24),
            ('Department',            'C', BLUE_LIGHT, BLUE_DARK, 16),
            ('Current Daily\nRate ₹', 'D', BLUE_LIGHT, BLUE_DARK, 16),
            ('NEW Daily\nRate ₹',     'E', GREEN_DARK, WHITE, 16),
            ('Effective From\n(YYYY-MM-DD)', 'F', GREEN_DARK, WHITE, 18),
            ('Hourly Rate\n(auto)',    'G', BLUE_LIGHT, BLUE_DARK, 14),
        ]
        for i, (label, col, bg, fg, width) in enumerate(headers, start=1):
            c = ws.cell(row=4, column=i, value=label)
            c.fill = hdr(bg)
            c.font = Font(color=fg, bold=True, size=10)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[4].height = 30

        # ── Row 5: instruction sub-row ────────────────────────────────────
        instructions = [
            'Do not edit',
            'Do not edit',
            'Do not edit',
            'Current active rate',
            '← FILL THIS COLUMN',
            'Optional — defaults to today',
            'Auto-calculated (÷8)',
        ]
        inst_colors = [GREY_LIGHT, GREY_LIGHT, GREY_LIGHT, GREY_LIGHT, 'FFF2CC', 'FFF2CC', GREY_LIGHT]
        for i, (txt, bg) in enumerate(zip(instructions, inst_colors), start=1):
            c = ws.cell(row=5, column=i, value=txt)
            c.fill = hdr(bg)
            c.font = Font(color='7F7F7F', italic=True, size=8)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        ws.row_dimensions[5].height = 14

        # ── Data rows (from row 6) ────────────────────────────────────────
        for idx, w in enumerate(workers):
            row = 6 + idx
            is_even = idx % 2 == 0
            ro_bg = GREY_LIGHT if is_even else WHITE
            ed_bg = GREEN_LIGHT if is_even else WHITE

            # A: employee_code (read-only style)
            c = ws.cell(row=row, column=1, value=w.get('employee_code', ''))
            cell_style(c, ro_bg, center=True)

            # B: employee_name
            c = ws.cell(row=row, column=2, value=w.get('employee_name', ''))
            cell_style(c, ro_bg)

            # C: department
            c = ws.cell(row=row, column=3, value=w.get('department', ''))
            cell_style(c, ro_bg)

            # D: current daily rate
            cur = w.get('daily_rate')
            c = ws.cell(row=row, column=4, value=float(cur) if cur else None)
            c.fill = hdr(ro_bg)
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = border
            if cur:
                c.number_format = '₹#,##0.00'

            # E: new daily rate (editable — leave empty)
            c = ws.cell(row=row, column=5, value=None)
            c.fill = hdr(ed_bg)
            c.font = Font(size=10, bold=True)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = border
            c.number_format = '₹#,##0.00'

            # F: effective_from (editable — pre-fill with today)
            c = ws.cell(row=row, column=6, value=str(date.today()))
            c.fill = hdr(ed_bg)
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border

            # G: hourly rate formula (auto)
            c = ws.cell(row=row, column=7, value=f'=IF(E{row}="","",E{row}/8)')
            c.fill = hdr(ro_bg)
            c.font = Font(size=10, color='595959', italic=True)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = border
            c.number_format = '₹#,##0.00'

            ws.row_dimensions[row].height = 18

        # ── Data validation: New Daily Rate must be > 0 ───────────────────
        if workers:
            last_row = 5 + len(workers)
            rate_dv = DataValidation(
                type='decimal', operator='greaterThan', formula1='0', allow_blank=True
            )
            rate_dv.error = 'Daily rate must be a positive number (e.g. 450.00)'
            rate_dv.errorTitle = 'Invalid Rate'
            rate_dv.prompt = 'Enter the new daily rate in ₹ (leave blank to skip this employee)'
            rate_dv.promptTitle = 'New Daily Rate'
            ws.add_data_validation(rate_dv)
            rate_dv.add(f'E6:E{last_row}')

        # ── Freeze panes at row 6 (keep headers visible) ──────────────────
        ws.freeze_panes = 'A6'

        # ── Protect read-only columns (A-D, G) — allow editing only E & F ─
        ws.protection.sheet = True
        ws.protection.password = 'readonly'
        ws.protection.enable()
        # Unlock columns E (new rate) and F (effective from)
        from openpyxl.styles.protection import Protection
        unlocked = Protection(locked=False)
        if workers:
            for row in range(6, 6 + len(workers)):
                ws.cell(row=row, column=5).protection = unlocked
                ws.cell(row=row, column=6).protection = unlocked

        # Save
        tmp = tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        return send_file(
            tmp.name,
            as_attachment=True,
            download_name=f'factory_rate_template_{date.today().isoformat()}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        current_app.logger.error(f'Rate template error: {str(e)}')
        return error_response(f'Failed to generate template: {str(e)}', 500)


# ============================================
# STATISTICS & SUMMARY
# ============================================

@factory_payroll_bp.route('/rates/statistics', methods=['GET'])
@jwt_required
@company_required
@role_required("HR", "MANAGER")
def get_rate_statistics():
    """
    Get statistics about factory worker rates
    
    Returns:
    - Total factory workers
    - Workers with rates assigned
    - Workers without rates
    - Average daily rate
    - Min/Max rates
    """
    try:
        result = FactoryPayrollService.get_all_workers_with_rates()
        
        if not result['success']:
            return error_response(result['message'], 400)
        
        workers = result.get('data', [])
        
        # Calculate statistics
        total_workers = len(workers)
        workers_with_rates = sum(1 for w in workers if w.get('daily_rate') is not None)
        workers_without_rates = total_workers - workers_with_rates
        
        rates = [w['daily_rate'] for w in workers if w.get('daily_rate') is not None]
        
        statistics = {
            'total_workers': total_workers,
            'workers_with_rates': workers_with_rates,
            'workers_without_rates': workers_without_rates,
            'average_daily_rate': round(sum(rates) / len(rates), 2) if rates else 0,
            'min_daily_rate': min(rates) if rates else 0,
            'max_daily_rate': max(rates) if rates else 0,
            'average_hourly_rate': round(sum(rates) / len(rates) / 8, 2) if rates else 0
        }
        
        return success_response('Statistics retrieved successfully', statistics)
        
    except Exception as e:
        return error_response(f"Error calculating statistics: {str(e)}", 500)


# ============================================
# PAYROLL CONFIGURATION ENDPOINTS
# ============================================

@factory_payroll_bp.route('/config', methods=['GET'])
@jwt_required
@company_required
@role_required("HR", "MANAGER")
def get_payroll_config():
    """Get current payroll configuration settings"""
    try:
        result = FactoryPayrollService.get_payroll_config()
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
            
    except Exception as e:
        return error_response(f"Error retrieving configuration: {str(e)}", 500)


@factory_payroll_bp.route('/config', methods=['PUT'])
@jwt_required
@company_required
@role_required("HR")
def update_payroll_config():
    """
    Update payroll configuration settings (Admin only)
    
    Request Body:
        {
            "full_day_hours": 12.0,
            "half_day_minimum_hours": 6.0,
            "absent_threshold_hours": 6.0,
            "hourly_divisor": 8,
            "overtime_multiplier": 2.0,
            "sunday_bonus_hours": 4.0
        }
    """
    try:
        data = request.get_json()
        
        if not data:
            return error_response("Request body is required", 400)
        
        # Get current user ID
        updated_by = get_jwt_identity()
        
        result = FactoryPayrollService.update_payroll_config(
            full_day_hours=data.get('full_day_hours'),
            half_day_minimum_hours=data.get('half_day_minimum_hours'),
            absent_threshold_hours=data.get('absent_threshold_hours'),
            hourly_divisor=data.get('hourly_divisor'),
            overtime_multiplier=data.get('overtime_multiplier'),
            sunday_bonus_hours=data.get('sunday_bonus_hours'),
            updated_by=updated_by
        )
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
            
    except Exception as e:
        return error_response(f"Error updating configuration: {str(e)}", 500)


# ============================================
# PAYROLL PERIOD ENDPOINTS
# ============================================

@factory_payroll_bp.route('/periods', methods=['GET'])
@jwt_required
@company_required
@role_required("HR", "MANAGER")
def get_payroll_periods():
    """Get all factory payroll periods"""
    try:
        result = FactoryPayrollService.get_payroll_periods()
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
            
    except Exception as e:
        return error_response(f"Error retrieving periods: {str(e)}", 500)


@factory_payroll_bp.route('/periods', methods=['POST'])
@jwt_required
@company_required
@role_required("HR")
def create_payroll_period():
    """
    Create a new factory payroll period
    
    Request Body:
        {
            "year": 2026,
            "month": 4
        }
    """
    try:
        data = request.get_json()
        
        if not data:
            return error_response("Request body is required", 400)
        
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return error_response("year and month are required", 400)
        
        # Get current user ID
        created_by = get_jwt_identity()
        
        result = FactoryPayrollService.create_payroll_period(
            year=year,
            month=month,
            created_by=created_by
        )
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
            
    except Exception as e:
        return error_response(f"Error creating period: {str(e)}", 500)


# ============================================
# PAYROLL CALCULATION ENDPOINTS
# ============================================

@factory_payroll_bp.route('/calculate', methods=['POST'])
@jwt_required
@company_required
@role_required("HR")
def calculate_payroll():
    """
    Calculate payroll for a period
    
    Request Body:
        {
            "period_id": 1
        }
    """
    try:
        data = request.get_json()
        
        if not data:
            return error_response("Request body is required", 400)
        
        period_id = data.get('period_id')
        
        if not period_id:
            return error_response("period_id is required", 400)
        
        # Get current user ID
        calculated_by = get_jwt_identity()
        
        result = FactoryPayrollService.calculate_payroll(
            period_id=period_id,
            calculated_by=calculated_by
        )
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
            
    except Exception as e:
        return error_response(f"Error calculating payroll: {str(e)}", 500)


@factory_payroll_bp.route('/summary/<int:period_id>', methods=['GET'])
@jwt_required
@company_required
@role_required("HR", "MANAGER")
def get_payroll_summary(period_id):
    """Get payroll summary for a specific period"""
    try:
        result = FactoryPayrollService.get_payroll_summary(period_id)
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
    except Exception as e:
        return error_response(f"Error retrieving payroll summary: {str(e)}", 500)


@factory_payroll_bp.route('/summary-by-month', methods=['GET'])
@jwt_required
@company_required
@role_required("HR", "MANAGER")
def get_payroll_summary_by_month():
    """Get payroll summary for a specific year/month (if period exists)"""
    try:
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        if not year or not month:
            return error_response("year and month are required", 400)
        result = FactoryPayrollService.get_payroll_summary_by_month(year, month)
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
    except Exception as e:
        return error_response(f"Error retrieving payroll summary: {str(e)}", 500)
    """Get payroll summary for a period"""
    try:
        result = FactoryPayrollService.get_payroll_summary(period_id)
        
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
            
    except Exception as e:
        return error_response(f"Error retrieving summary: {str(e)}", 500)


@factory_payroll_bp.route('/lock', methods=['POST'])
@jwt_required
@company_required
@role_required("HR")
def lock_payroll_period():
    """Lock a payroll period"""
    try:
        data = request.get_json()
        if not data or not data.get('period_id'):
            return error_response("period_id is required", 400)
        locked_by = get_jwt_identity()
        result = FactoryPayrollService.lock_payroll_period(
            period_id=data['period_id'],
            locked_by=locked_by
        )
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
    except Exception as e:
        return error_response(f"Error locking period: {str(e)}", 500)


@factory_payroll_bp.route('/unlock', methods=['POST'])
@jwt_required
@company_required
@role_required("HR")
def unlock_payroll_period():
    """Unlock a locked payroll period (revert to CALCULATED)"""
    try:
        data = request.get_json()
        if not data or not data.get('period_id'):
            return error_response("period_id is required", 400)
        unlocked_by = get_jwt_identity()
        result = FactoryPayrollService.unlock_payroll_period(
            period_id=data['period_id'],
            unlocked_by=unlocked_by
        )
        if result['success']:
            return success_response(result['message'], result.get('data'))
        else:
            return error_response(result['message'], 400)
    except Exception as e:
        return error_response(f"Error unlocking period: {str(e)}", 500)


@factory_payroll_bp.route('/periods/<int:period_id>', methods=['DELETE'])
@jwt_required
@company_required
@role_required("HR")
def delete_payroll_period(period_id):
    """Delete a payroll period (DRAFT or CALCULATED only, not LOCKED)"""
    try:
        result = FactoryPayrollService.delete_payroll_period(period_id)
        if result['success']:
            return success_response(result['message'])
        else:
            return error_response(result['message'], 400)
    except Exception as e:
        return error_response(f"Error deleting period: {str(e)}", 500)
